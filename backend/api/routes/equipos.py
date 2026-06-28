from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from database import get_session
from models.equipos import Equipo
from models.estados import EstadoEquipo
from models.users import Usuario
from models.ordenes import OrdenTrabajo
from models.documentos import DocumentoAdjunto
from models.preventivo import TareaPreventiva
from models.historial import EventoHistorial
from models.proveedores import Proveedor
from schemas.equipo import EquipoCreate, EquipoRead, EquipoUpdate
from schemas.proveedor import ProveedorCreate, ProveedorRead
from datetime import date, datetime
from io import BytesIO, StringIO
import openpyxl
import csv
import os
import uuid
from pathlib import Path
from config import get_dir, sanitize_filename
from utils.meta_json import write_meta_json, build_equipo_meta

router = APIRouter(prefix="/equipos", tags=["Equipos"])


# ============================================================
# RUTAS ESTÁTICAS (deben ir ANTES de /{equipo_id})
# ============================================================

# --- Endpoint para listar estados de equipo (catálogo) ---
@router.get("/estados")
def listar_estados(session: Session = Depends(get_session)):
    estados = session.exec(select(EstadoEquipo)).all()
    return estados


# --- Endpoint para crear estado (solo administración) ---
@router.post("/estados")
def crear_estado(nombre_estado: str, session: Session = Depends(get_session)):
    existe = session.exec(select(EstadoEquipo).where(EstadoEquipo.nombre_estado == nombre_estado)).first()
    if existe:
        raise HTTPException(status_code=400, detail="El estado ya existe")
    nuevo_estado = EstadoEquipo(nombre_estado=nombre_estado)
    session.add(nuevo_estado)
    session.commit()
    session.refresh(nuevo_estado)
    return nuevo_estado


# --- Endpoint para listar proveedores (para dropdown en frontend) ---
@router.get("/proveedores")
def listar_proveedores_para_dropdown(session: Session = Depends(get_session)):
    """Lista proveedores para llenar el dropdown de proveedor_principal_id en el formulario de equipo."""
    proveedores = session.exec(select(Proveedor).order_by(Proveedor.nombre_empresa)).all()
    return [{"id": p.id, "nombre_empresa": p.nombre_empresa, "ciudad": p.ciudad} for p in proveedores]


# --- Endpoint para crear proveedor al vuelo desde el formulario de equipo ---
@router.post("/from-proveedor-nombre", response_model=ProveedorRead, status_code=201)
def crear_proveedor_desde_nombre(payload: dict, session: Session = Depends(get_session)):
    """
    Crea un proveedor con solo el nombre_empresa (los demás datos se completan después
    desde la página de Proveedores). Útil cuando el usuario registra un equipo de un
    proveedor que aún no está en el directorio.

    Payload esperado: {"nombre_empresa": "TechMed Bolivia SRL"}
    """
    nombre = payload.get("nombre_empresa", "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="nombre_empresa es obligatorio")

    existe = session.exec(
        select(Proveedor).where(Proveedor.nombre_empresa == nombre)
    ).first()
    if existe:
        raise HTTPException(status_code=400, detail=f"Ya existe un proveedor con el nombre '{nombre}'")

    nuevo = Proveedor(nombre_empresa=nombre)
    session.add(nuevo)
    session.commit()
    session.refresh(nuevo)
    return nuevo


# ============================================================
# CRUD PRINCIPAL DE EQUIPOS
# ============================================================

@router.post("/", response_model=EquipoRead)
def crear_equipo(equipo: EquipoCreate, session: Session = Depends(get_session)):
    existe = session.exec(select(Equipo).where(Equipo.numero_serie == equipo.numero_serie)).first()
    if existe:
        raise HTTPException(status_code=400, detail="El número de serie ya está registrado")

    # Validar que proveedor_principal_id (si viene) exista
    if equipo.proveedor_principal_id:
        prov = session.get(Proveedor, equipo.proveedor_principal_id)
        if not prov:
            raise HTTPException(status_code=400, detail=f"Proveedor ID {equipo.proveedor_principal_id} no existe")

    db_equipo = Equipo(**equipo.model_dump())
    session.add(db_equipo)
    session.commit()
    session.refresh(db_equipo)

    # Crear carpeta del equipo y escribir .meta.json
    try:
        equipo_code = f"E{db_equipo.id:04d}"
        modelo_safe = sanitize_filename(db_equipo.modelo, "SM")
        serie_safe = sanitize_filename(db_equipo.numero_serie, "SN")
        folder_name = f"{equipo_code}_{modelo_safe}_{serie_safe}"
        equipo_dir = get_dir("equipos_imagenes") / folder_name
        equipo_dir.mkdir(parents=True, exist_ok=True)
        meta_data = build_equipo_meta(db_equipo)
        write_meta_json(equipo_dir, meta_data)
    except Exception as e:
        print(f"[equipos.py] WARNING: No se pudo crear .meta.json: {e}")

    return db_equipo


@router.get("/")
def listar_equipos(session: Session = Depends(get_session)):
    """
    Lista todos los equipos con campos calculados:
    - en_garantia: bool (calculado desde fecha_inicio_garantia y fecha_fin_garantia)
    - en_contrato: bool (calculado consultando ContratoEquipo + Contrato vigente)
    """
    from models.contratos import Contrato, ContratoEquipo
    from datetime import date as date_type

    equipos = session.exec(select(Equipo)).all()
    hoy = date_type.today()

    # Pre-cargar todos los ContratoEquipo para evitar N+1 queries
    todas_asociaciones = session.exec(select(ContratoEquipo)).all()
    todos_contratos = {c.id: c for c in session.exec(select(Contrato)).all()}

    resultado = []
    for eq in equipos:
        data = eq.model_dump()

        # Calcular en_garantia
        en_garantia = False
        if eq.fecha_inicio_garantia and eq.fecha_fin_garantia:
            inicio = eq.fecha_inicio_garantia
            fin = eq.fecha_fin_garantia
            if hasattr(inicio, 'date'):
                inicio = inicio.date()
            if hasattr(fin, 'date'):
                fin = fin.date()
            if inicio <= hoy <= fin:
                en_garantia = True
        data['en_garantia'] = en_garantia

        # Calcular en_contrato
        en_contrato = False
        equipo_contrato_ids = [a.contrato_id for a in todas_asociaciones if a.equipo_id == eq.id]
        for contrato_id in equipo_contrato_ids:
            contrato = todos_contratos.get(contrato_id)
            if contrato:
                inicio = contrato.fecha_inicio
                fin = contrato.fecha_fin
                if hasattr(inicio, 'date'):
                    inicio = inicio.date()
                if hasattr(fin, 'date'):
                    fin = fin.date()
                if inicio <= hoy <= fin:
                    en_contrato = True
                    break
        data['en_contrato'] = en_contrato

        resultado.append(data)

    return resultado


@router.put("/{equipo_id}", response_model=EquipoRead)
def actualizar_equipo(equipo_id: int, equipo_data: EquipoUpdate, session: Session = Depends(get_session)):
    """
    Actualiza un equipo. Campos NO editables después de creado:
    - modelo, marca, numero_serie (afectan carpetas y .meta.json)

    Si el frontend intenta cambiar estos campos, se rechaza con error 400.
    """
    db_equipo = session.get(Equipo, equipo_id)
    if not db_equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")

    equipo_data_dict = equipo_data.model_dump(exclude_unset=True)

    # Verificar que no se estén intentando cambiar campos no editables
    campos_bloqueados = []
    if "modelo" in equipo_data_dict and equipo_data_dict["modelo"] != db_equipo.modelo:
        campos_bloqueados.append("modelo")
    if "marca" in equipo_data_dict and equipo_data_dict["marca"] != db_equipo.marca:
        campos_bloqueados.append("marca")
    if "numero_serie" in equipo_data_dict and equipo_data_dict["numero_serie"] != db_equipo.numero_serie:
        campos_bloqueados.append("numero_serie")

    if campos_bloqueados:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Los siguientes campos no son modificables después de creado el equipo: "
                f"{', '.join(campos_bloqueados)}. Esto es porque afectan la estructura de "
                f"carpetas y archivos del sistema. Si necesita corregir un error, "
                f"elimine el equipo y créelo de nuevo."
            )
        )

    # Validar proveedor_principal_id si viene
    if "proveedor_principal_id" in equipo_data_dict and equipo_data_dict["proveedor_principal_id"]:
        prov = session.get(Proveedor, equipo_data_dict["proveedor_principal_id"])
        if not prov:
            raise HTTPException(
                status_code=400,
                detail=f"Proveedor ID {equipo_data_dict['proveedor_principal_id']} no existe"
            )

    for key, value in equipo_data_dict.items():
        setattr(db_equipo, key, value)

    session.add(db_equipo)
    session.commit()
    session.refresh(db_equipo)

    # Actualizar .meta.json
    try:
        equipo_code = f"E{db_equipo.id:04d}"
        modelo_safe = sanitize_filename(db_equipo.modelo, "SM")
        serie_safe = sanitize_filename(db_equipo.numero_serie, "SN")
        folder_name = f"{equipo_code}_{modelo_safe}_{serie_safe}"
        equipo_dir = get_dir("equipos_imagenes") / folder_name
        if equipo_dir.exists():
            meta_data = build_equipo_meta(db_equipo, db_equipo.imagen_ruta)
            write_meta_json(equipo_dir, meta_data)
    except Exception as e:
        print(f"[equipos.py] WARNING: No se pudo actualizar .meta.json: {e}")

    return db_equipo


@router.delete("/{equipo_id}")
def eliminar_equipo(equipo_id: int, session: Session = Depends(get_session)):
    """
    Elimina un equipo SOLO si no tiene dependencias.

    Dependencias que bloquean la eliminación:
    - Órdenes de Trabajo (OrdenTrabajo.equipo_id)
    - Documentos adjuntos (DocumentoAdjunto.equipo_id)
    - Tareas preventivas (TareaPreventiva.equipo_id)
    - Eventos de historial (EventoHistorial.equipo_id)

    Si el usuario quiere "retirar" el equipo sin eliminarlo, debe cambiar
    el estado_id a "Retirado/Baja" (id=12 en el catálogo).
    """
    db_equipo = session.get(Equipo, equipo_id)
    if not db_equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")

    # Verificar dependencias
    ots = session.exec(
        select(OrdenTrabajo).where(OrdenTrabajo.equipo_id == equipo_id)
    ).all()
    documentos = session.exec(
        select(DocumentoAdjunto).where(DocumentoAdjunto.equipo_id == equipo_id)
    ).all()
    mps = session.exec(
        select(TareaPreventiva).where(TareaPreventiva.equipo_id == equipo_id)
    ).all()
    historial = session.exec(
        select(EventoHistorial).where(EventoHistorial.equipo_id == equipo_id)
    ).all()

    dependencias = []
    if ots:
        dependencias.append(f"{len(ots)} orden(s) de trabajo")
    if documentos:
        dependencias.append(f"{len(documentos)} documento(s)")
    if mps:
        dependencias.append(f"{len(mps)} tarea(s) preventiva(s)")
    if historial:
        dependencias.append(f"{len(historial)} evento(s) de historial")

    if dependencias:
        raise HTTPException(
            status_code=400,
            detail=(
                f"No se puede eliminar el equipo porque tiene dependencias asociadas: "
                f"{', '.join(dependencias)}. "
                f"Si desea retirar el equipo del servicio sin eliminarlo, cambie su estado "
                f"a 'Retirado/Baja' desde la página de Equipos."
            )
        )

    # Sin dependencias: eliminar carpeta física + registro BD
    try:
        equipo_code = f"E{db_equipo.id:04d}"
        modelo_safe = sanitize_filename(db_equipo.modelo, "SM")
        serie_safe = sanitize_filename(db_equipo.numero_serie, "SN")
        folder_name = f"{equipo_code}_{modelo_safe}_{serie_safe}"
        equipo_dir = get_dir("equipos_imagenes") / folder_name
        if equipo_dir.exists():
            import shutil
            shutil.rmtree(equipo_dir)
    except Exception as e:
        print(f"[equipos.py] WARNING: No se pudo eliminar carpeta del equipo: {e}")

    session.delete(db_equipo)
    session.commit()
    return {"ok": True, "message": "Equipo eliminado correctamente"}

# ---------------------------------------------------------
# ENDPOINT: SUBIR IMAGEN PRINCIPAL DE EQUIPO
# Estructura: uploads/EQUIPOS/EXXXX_numero_serie/EXXXX_numero_serie.ext
# ---------------------------------------------------------
@router.post("/{equipo_id}/upload_imagen")
async def upload_imagen_equipo(equipo_id: int, file: UploadFile = File(...), session: Session = Depends(get_session)):
    """
    Sube una imagen principal para un equipo.
    Guarda en uploads/EQUIPOS/EXXXX_Modelo_serie/EXXXX_Modelo_serie.ext
    """
    db_equipo = session.get(Equipo, equipo_id)
    if not db_equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    
    # Validar tipo de archivo
    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp", "image/bmp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Use JPEG, PNG, GIF, WebP o BMP.")
    
    # Validar tamaño (5MB máximo)
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="La imagen no debe superar 5MB")
    
    # Generar código de equipo: E0001, E0002, etc.
    equipo_code = f"E{equipo_id:04d}"
    # Usar modelo y numero_serie para el nombre de carpeta (sanitizar)
    modelo_safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in (db_equipo.modelo or "SM"))
    serie_safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in (db_equipo.numero_serie or "SN"))
    folder_name = f"{equipo_code}_{modelo_safe}_{serie_safe}"
    
    # Crear carpeta: uploads/EQUIPOS/EXXXX_serie/
    uploads_dir = get_dir("equipos_imagenes") / folder_name
    uploads_dir.mkdir(parents=True, exist_ok=True)
    
    # Nombre del archivo: EXXXX_serie.ext (misma extensión original)
    ext = os.path.splitext(file.filename)[1] if file.filename else ".jpg"
    filename = f"{folder_name}{ext}"
    file_path = uploads_dir / filename
    
    # Si ya existe una imagen previa con otra extensión, eliminarla
    for old_ext in [".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"]:
        old_path = uploads_dir / f"{folder_name}{old_ext}"
        if old_path.exists() and old_path != file_path:
            old_path.unlink()
    
    # Guardar archivo
    with open(str(file_path), "wb") as f:
        f.write(contents)
    
    # Actualizar imagen_ruta en la BD (ruta relativa para servir vía /uploads/)
    imagen_ruta = f"EQUIPOS/{folder_name}/{filename}"
    db_equipo.imagen_ruta = imagen_ruta
    session.add(db_equipo)
    session.commit()
    session.refresh(db_equipo)
    
    # Actualizar .meta.json con la nueva imagen
    try:
        meta_data = build_equipo_meta(db_equipo, imagen_ruta)
        write_meta_json(uploads_dir, meta_data)
    except Exception as e:
        print(f"[equipos.py] WARNING: No se pudo actualizar .meta.json: {e}")
    
    return {"ok": True, "imagen_ruta": db_equipo.imagen_ruta, "nombre_archivo": file.filename}


# ---------------------------------------------------------
# ENDPOINT: ELIMINAR IMAGEN PRINCIPAL DE EQUIPO
# ---------------------------------------------------------
@router.delete("/{equipo_id}/imagen")
def eliminar_imagen_equipo(equipo_id: int, session: Session = Depends(get_session)):
    """Elimina la imagen principal de un equipo (archivo + BD)."""
    db_equipo = session.get(Equipo, equipo_id)
    if not db_equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    
    if not db_equipo.imagen_ruta:
        raise HTTPException(status_code=400, detail="Este equipo no tiene imagen")
    
    # Eliminar archivo físico
    file_path = get_dir("uploads_base") / db_equipo.imagen_ruta
    if file_path.exists():
        file_path.unlink()
    
    # Limpiar BD
    db_equipo.imagen_ruta = None
    session.add(db_equipo)
    session.commit()
    
    # Actualizar .meta.json (sin imagen)
    try:
        equipo_code = f"E{db_equipo.id:04d}"
        modelo_safe = sanitize_filename(db_equipo.modelo, "SM")
        serie_safe = sanitize_filename(db_equipo.numero_serie, "SN")
        folder_name = f"{equipo_code}_{modelo_safe}_{serie_safe}"
        equipo_dir = get_dir("equipos_imagenes") / folder_name
        if equipo_dir.exists():
            meta_data = build_equipo_meta(db_equipo)
            write_meta_json(equipo_dir, meta_data)
    except Exception as e:
        print(f"[equipos.py] WARNING: No se pudo actualizar .meta.json: {e}")
    
    return {"ok": True, "message": "Imagen eliminada"}

# ---------------------------------------------------------
# ENDPOINT: IMPORTAR EQUIPOS DESDE EXCEL
# ---------------------------------------------------------
@router.post("/import-excel")
async def importar_equipos_excel(file: UploadFile = File(...), session: Session = Depends(get_session)):
    """
    Importa equipos desde un archivo Excel (.xlsx) o CSV - v0.9.0.

    Columnas esperadas (encabezado en fila 1):
      nombre_corto*, modelo*, numero_serie*, numero_material, marca*,
      fecha_adquisicion, fecha_inicio_garantia, fecha_fin_garantia,
      ubicacion_actual, estado, proveedor_principal, condicion_origen,
      descripcion, observaciones

    Cambios v0.9.0:
    - OBLIGATORIAS: nombre_corto, modelo, numero_serie, marca (fecha_adquisicion ahora es opcional)
    - ELIMINADAS: registro_sanitario_bolivia, calibracion_proxima, responsable_username
    - NUEVAS: fecha_inicio_garantia, condicion_origen, observaciones
    - proveedor_principal (texto): si no existe el proveedor en BD, se CREA automáticamente
    - estado: se resuelve por nombre (debe existir en EstadoEquipo, default=1 Operativo)
    - Si numero_serie ya existe, se ACTUALIZA el registro (upsert)
    """
    # Validar extensión
    ext = Path(file.filename).suffix.lower() if file.filename else ''
    if ext not in ('.xlsx', '.csv'):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .csv")

    # Leer el archivo
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:  # 5MB límite
        raise HTTPException(status_code=400, detail="El archivo no debe superar 5MB")

    # Determinar si es Excel o CSV
    is_csv = ext == '.csv'

    if is_csv:
        try:
            text = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                text = contents.decode('latin-1')
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error al decodificar el archivo CSV: {str(e)}")

        try:
            sniffer_sample = text[:2048]
            try:
                dialect = csv.Sniffer().sniff(sniffer_sample, delimiters=',;\t')
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(StringIO(text), dialect)
            filas = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo CSV: {str(e)}")

        if len(filas) < 2:
            raise HTTPException(status_code=400, detail="El archivo CSV está vacío o solo tiene encabezados")
        ws = None
    else:
        try:
            wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo Excel: {str(e)}")

    # Obtener estados para resolver nombres
    estados_db = session.exec(select(EstadoEquipo)).all()
    estado_map = {e.nombre_estado.strip().lower(): e for e in estados_db}

    # Columnas esperadas v0.9.0 (sin campos obsoletos, con campos nuevos)
    COLUMNAS = [
        'nombre_corto', 'modelo', 'numero_serie', 'numero_material', 'marca',
        'fecha_adquisicion', 'fecha_inicio_garantia', 'fecha_fin_garantia',
        'ubicacion_actual', 'estado', 'proveedor_principal', 'condicion_origen',
        'descripcion', 'observaciones'
    ]
    OBLIGATORIAS = {'nombre_corto', 'modelo', 'numero_serie', 'marca'}

    # Valores válidos para condicion_origen
    CONDICIONES_VALIDAS = {
        'Compra', 'Donación', 'Préstamo', 'Demostración', 'Evaluación',
        'Leasing', 'Renta', 'Comodato', 'Otro'
    }

    # Buscar la hoja con encabezados (solo Excel)
    if not is_csv:
        ws = None
        for sheet in wb.worksheets:
            filas_tmp = list(sheet.iter_rows(values_only=True))
            if filas_tmp:
                enc_tmp = [str(c).strip().lower() if c else '' for c in filas_tmp[0]]
                if OBLIGATORIAS.issubset(set(enc_tmp)):
                    ws = sheet
                    break
        if ws is None:
            ws = wb.worksheets[0]
        filas = list(ws.iter_rows(values_only=True))
        if len(filas) < 2:
            raise HTTPException(status_code=400, detail="El archivo está vacío o solo tiene encabezados")

    encabezados_leidos = [str(c).strip().lower() if c else '' for c in filas[0]]
    col_index = {}
    for i, h in enumerate(encabezados_leidos):
        if h in COLUMNAS:
            col_index[h] = i

    faltantes = OBLIGATORIAS - set(col_index.keys())
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas obligatorias: {', '.join(faltantes)}. Encabezados encontrados: {', '.join(encabezados_leidos)}"
        )

    exitosos = 0
    actualizados = 0
    fallidos = []
    proveedores_creados = 0  # contador de proveedores creados al vuelo

    for fila_num, fila in enumerate(filas[1:], start=2):
        try:
            def get_val(col_name):
                idx = col_index.get(col_name)
                if idx is None or idx >= len(fila):
                    return None
                return fila[idx]

            def get_str(col_name):
                val = get_val(col_name)
                if val is None:
                    return None
                return str(val).strip()

            # Validar obligatorios
            errores_fila = []
            for col in OBLIGATORIAS:
                val = get_str(col)
                if not val:
                    errores_fila.append(f"'{col}' es obligatorio")
            if errores_fila:
                fallidos.append({
                    "fila": fila_num,
                    "numero_serie": get_str('numero_serie') or 'N/A',
                    "errores": errores_fila
                })
                continue

            # Parsear fechas (todas opcionales en v0.9.0)
            fecha_adq = None
            adq_raw = get_val('fecha_adquisicion')
            if adq_raw and str(adq_raw).strip() not in ('', 'None'):
                try:
                    fecha_adq = _parse_date(adq_raw)
                except ValueError:
                    fallidos.append({
                        "fila": fila_num,
                        "numero_serie": get_str('numero_serie') or 'N/A',
                        "errores": [f"fecha_adquisicion inválida: '{adq_raw}'. Use YYYY-MM-DD"]
                    })
                    continue

            fecha_inicio_gar = None
            fig_raw = get_val('fecha_inicio_garantia')
            if fig_raw:
                try:
                    fecha_inicio_gar = _parse_date(fig_raw)
                except ValueError:
                    pass

            fecha_fin_gar = None
            ffg_raw = get_val('fecha_fin_garantia')
            if ffg_raw:
                try:
                    fecha_fin_gar = _parse_date(ffg_raw)
                except ValueError:
                    pass

            # Validar que fecha_inicio_garantia <= fecha_fin_garantia
            if fecha_inicio_gar and fecha_fin_gar and fecha_inicio_gar > fecha_fin_gar:
                fallidos.append({
                    "fila": fila_num,
                    "numero_serie": get_str('numero_serie') or 'N/A',
                    "errores": ["fecha_inicio_garantia debe ser <= fecha_fin_garantia"]
                })
                continue

            # Resolver estado_id
            estado_id = 1  # Default: Operativo
            estado_str = get_str('estado')
            if estado_str:
                estado_key = estado_str.strip().lower()
                if estado_key in estado_map:
                    estado_id = estado_map[estado_key].id
                else:
                    errores_fila.append(f"Estado '{estado_str}' no encontrado, se usará el default (Operativo)")

            # Validar condicion_origen
            cond_origen = get_str('condicion_origen')
            if cond_origen and cond_origen not in CONDICIONES_VALIDAS:
                fallidos.append({
                    "fila": fila_num,
                    "numero_serie": get_str('numero_serie') or 'N/A',
                    "errores": [f"condicion_origen '{cond_origen}' inválido. Valores válidos: {', '.join(sorted(CONDICIONES_VALIDAS))}"]
                })
                continue

            # Resolver proveedor_principal (texto → FK, crear si no existe)
            proveedor_id = None
            prov_str = get_str('proveedor_principal')
            if prov_str:
                prov_existente = session.exec(
                    select(Proveedor).where(Proveedor.nombre_empresa == prov_str)
                ).first()
                if prov_existente:
                    proveedor_id = prov_existente.id
                else:
                    # Crear proveedor al vuelo (solo con nombre_empresa)
                    nuevo_prov = Proveedor(nombre_empresa=prov_str)
                    session.add(nuevo_prov)
                    session.flush()  # para obtener el ID sin commit
                    proveedor_id = nuevo_prov.id
                    proveedores_creados += 1

            # Verificar si ya existe (por numero_serie) - upsert
            num_serie = get_str('numero_serie')
            equipo_existente = session.exec(
                select(Equipo).where(Equipo.numero_serie == num_serie)
            ).first()

            if equipo_existente:
                # ACTUALIZAR (no tocar modelo/marca/numero_serie que son no editables)
                equipo_existente.nombre_corto = get_str('nombre_corto') or equipo_existente.nombre_corto
                equipo_existente.numero_material = get_str('numero_material') or equipo_existente.numero_material
                equipo_existente.fecha_adquisicion = fecha_adq or equipo_existente.fecha_adquisicion
                equipo_existente.fecha_inicio_garantia = fecha_inicio_gar or equipo_existente.fecha_inicio_garantia
                equipo_existente.fecha_fin_garantia = fecha_fin_gar or equipo_existente.fecha_fin_garantia
                equipo_existente.ubicacion_actual = get_str('ubicacion_actual') or equipo_existente.ubicacion_actual
                equipo_existente.estado_id = estado_id
                equipo_existente.proveedor_principal_id = proveedor_id or equipo_existente.proveedor_principal_id
                equipo_existente.condicion_origen = cond_origen or equipo_existente.condicion_origen
                equipo_existente.descripcion = get_str('descripcion') or equipo_existente.descripcion
                equipo_existente.observaciones = get_str('observaciones') or equipo_existente.observaciones
                session.add(equipo_existente)
                actualizados += 1
            else:
                # CREAR nuevo
                nuevo = Equipo(
                    nombre_corto=get_str('nombre_corto'),
                    modelo=get_str('modelo'),
                    numero_serie=num_serie,
                    numero_material=get_str('numero_material'),
                    marca=get_str('marca'),
                    fecha_adquisicion=fecha_adq,
                    fecha_inicio_garantia=fecha_inicio_gar,
                    fecha_fin_garantia=fecha_fin_gar,
                    ubicacion_actual=get_str('ubicacion_actual'),
                    estado_id=estado_id,
                    proveedor_principal_id=proveedor_id,
                    condicion_origen=cond_origen,
                    descripcion=get_str('descripcion'),
                    observaciones=get_str('observaciones'),
                )
                session.add(nuevo)
                exitosos += 1

        except Exception as e:
            fallidos.append({
                "fila": fila_num,
                "numero_serie": get_str('numero_serie') if col_index.get('numero_serie') is not None else 'N/A',
                "errores": [str(e)]
            })

    session.commit()
    if not is_csv:
        wb.close()

    return {
        "exitosos": exitosos,
        "actualizados": actualizados,
        "fallidos": len(fallidos),
        "total_procesados": exitosos + actualizados + len(fallidos),
        "proveedores_creados": proveedores_creados,
        "errores": fallidos
    }


def _parse_date(val) -> date:
    """Parsea una fecha en múltiples formatos.
    
    Acepta:
    - Objetos date o datetime de Python (de openpyxl)
    - Strings en formato YYYY-MM-DD, DD/MM/YYYY, YYYY/MM/DD, DD-MM-YYYY
    - Strings datetime de Excel como '2003-12-30 00:00:00'
    """
    # Si ya es un objeto date, devolverlo directamente
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    # Si es datetime (openpyxl devuelve datetime para celdas de fecha), extraer date
    if isinstance(val, datetime):
        return val.date()
    # Convertir a string si no lo es
    val = str(val).strip()
    # Intentar formatos datetime primero (Excel envía '2003-12-30 00:00:00')
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    # Intentar formatos solo fecha
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"No se pudo parsear la fecha: '{val}'")


# ---------------------------------------------------------
# ENDPOINT: DESCARGAR PLANTILLA CSV
# ---------------------------------------------------------
@router.get("/plantilla-csv")
def descargar_plantilla_csv(session: Session = Depends(get_session)):
    """
    [RESPALDO] La fuente principal de esta plantilla ahora es el archivo estatico en:
        frontend/public/plantillas/plantilla_equipos.csv
    El frontend descarga directamente desde ahi (sin llamar al backend).
    Este endpoint se mantiene como respaldo y para documentacion Swagger.

    v0.9.0: Plantilla actualizada con nuevos campos (sin campos obsoletos).
    """
    output = StringIO()
    encabezados = [
        'nombre_corto', 'modelo', 'numero_serie', 'numero_material', 'marca',
        'fecha_adquisicion', 'fecha_inicio_garantia', 'fecha_fin_garantia',
        'ubicacion_actual', 'estado', 'proveedor_principal', 'condicion_origen',
        'descripcion', 'observaciones'
    ]

    writer = csv.writer(output)
    writer.writerow(encabezados)

    # 2 filas de ejemplo (plantilla vacía según filosofía v0.9.0)
    datos_demo = [
        ["Microscopio Olympus CX23", "CX23", "MIC-OLY-001", "MAT-CX23-A", "Olympus",
         "2023-03-15", "2023-03-15", "2025-03-15", "Lab. Microbiología", "Operativo",
         "Olympus Bolivia", "Compra",
         "Microscopio binocular para microbiología clínica", "Funciona correctamente"],

        ["Monitor Signos Vitales Mindray", "uMEC10", "MON-MIN-002", "", "Mindray",
         "2022-11-05", "2022-11-05", "2024-11-05", "UCI Box 3", "En Mantenimiento",
         "Mindray Bolivia", "Donación",
         "Monitor multiparámetro con SpO2, ECG, NIBP, Temp", "Requiere calibración de SpO2"],
    ]

    for fila in datos_demo:
        writer.writerow(fila)

    output.seek(0)
    from datetime import datetime as dt
    filename = f"CMMS-BioAI_Plantilla_Equipos_{dt.now().strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "text/csv; charset=utf-8"
        }
    )


# ---------------------------------------------------------
# ENDPOINT: DESCARGAR PLANTILLA EXCEL DEMO
# ---------------------------------------------------------
@router.get("/plantilla-excel")
def descargar_plantilla_excel(session: Session = Depends(get_session)):
    """
    [RESPALDO] La fuente principal de esta plantilla ahora es el archivo estatico en:
        frontend/public/plantillas/plantilla_equipos.xlsx
    El frontend descarga directamente desde ahi (sin llamar al backend).
    Este endpoint se mantiene como respaldo y para documentacion Swagger.

    v0.9.0: Plantilla actualizada con nuevos campos (sin campos obsoletos).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipos CMMS-BioAI"

    # Encabezados v0.9.0
    encabezados = [
        'nombre_corto', 'modelo', 'numero_serie', 'numero_material', 'marca',
        'fecha_adquisicion', 'fecha_inicio_garantia', 'fecha_fin_garantia',
        'ubicacion_actual', 'estado', 'proveedor_principal', 'condicion_origen',
        'descripcion', 'observaciones'
    ]
    ws.append(encabezados)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, _ in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 2 filas de ejemplo (plantilla vacía según filosofía v0.9.0)
    datos_demo = [
        ["Microscopio Olympus CX23", "CX23", "MIC-OLY-001", "MAT-CX23-A", "Olympus",
         "2023-03-15", "2023-03-15", "2025-03-15", "Lab. Microbiología", "Operativo",
         "Olympus Bolivia", "Compra",
         "Microscopio binocular para microbiología clínica", "Funciona correctamente"],

        ["Monitor Signos Vitales Mindray", "uMEC10", "MON-MIN-002", "", "Mindray",
         "2022-11-05", "2022-11-05", "2024-11-05", "UCI Box 3", "En Mantenimiento",
         "Mindray Bolivia", "Donación",
         "Monitor multiparámetro con SpO2, ECG, NIBP, Temp", "Requiere calibración de SpO2"],
    ]

    data_align = Alignment(vertical="center", wrap_text=True)
    for fila_data in datos_demo:
        ws.append(fila_data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = data_align

    anchos = [28, 18, 18, 16, 18, 18, 18, 18, 24, 18, 24, 18, 40, 35]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    # Hoja de instrucciones v0.9.0
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA IMPORTAR EQUIPOS v0.9.0"],
        [""],
        ["1. Campos obligatorios (*) no pueden estar vacíos:"],
        ["   - nombre_corto *", "nombre descriptivo del equipo (alias)"],
        ["   - modelo *", "modelo del equipo (NO editable después de creado)"],
        ["   - numero_serie *", "único, si ya existe se ACTUALIZA el registro (upsert)"],
        ["   - marca *", "fabricante (NO editable después de creado)"],
        [""],
        ["2. Campo 'estado': debe coincidir con un estado existente en el sistema."],
        ["   Estados típicos: Operativo, En Mantenimiento, Fuera de Servicio, etc."],
        ["   Si no coincide, se usará el estado por defecto (Operativo)."],
        [""],
        ["3. Campo 'proveedor_principal': nombre del proveedor."],
        ["   - Si el proveedor NO existe en la BD, se CREA automáticamente"],
        ["   - Luego podrás completar sus datos (ciudad, contacto, etc.) en la página Proveedores"],
        [""],
        ["4. Campo 'condicion_origen' (valores permitidos):"],
        ["   Compra, Donación, Préstamo, Demostración, Evaluación, Leasing, Renta, Comodato, Otro"],
        [""],
        ["5. Fechas: formato YYYY-MM-DD o DD/MM/YYYY."],
        ["   - fecha_adquisicion: opcional"],
        ["   - fecha_inicio_garantia: opcional (debe ser <= fecha_fin_garantia)"],
        ["   - fecha_fin_garantia: opcional"],
        [""],
        ["6. Campos 'descripcion' y 'observaciones':"],
        ["   - descripcion: descripción TÉCNICA del equipo (¿qué es? ¿qué hace?)"],
        ["   - observaciones: notas OPERATIVAS (¿cómo está? accesorios, advertencias)"],
        [""],
        ["7. Límite de archivo: 5MB. Solo archivos .xlsx o .csv."],
        [""],
        ["8. CAMPOS ELIMINADOS en v0.9.0 (no incluirlos):"],
        ["   - registro_sanitario_bolivia (universalidad)"],
        ["   - calibracion_proxima (gestionado vía MP/OT)"],
        ["   - responsable_username (asignación flexible en OT/MP)"],
        [""],
        ["9. Puede eliminar las filas de ejemplo y usar sus propios datos."],
        ["   Mantenga los encabezados de columna en la fila 1."],
    ]
    for row in instrucciones:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 55
    ws2.column_dimensions['B'].width = 60

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()

    from datetime import datetime as dt
    filename = f"CMMS-BioAI_Plantilla_Equipos_{dt.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )