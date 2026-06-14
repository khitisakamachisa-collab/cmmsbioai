from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from database import get_session
from models.herramientas import Herramienta
from schemas.herramienta import HerramientaCreate, HerramientaRead, HerramientaUpdate
from io import BytesIO, StringIO
from datetime import datetime as dt
from pathlib import Path
import openpyxl
import csv
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import get_dir, sanitize_filename
from utils.meta_json import write_meta_json, build_herramienta_meta

router = APIRouter(prefix="/herramientas", tags=["Herramientas"])

# ============================================================
# RUTAS ESTÁTICAS (deben ir ANTES de /{herr_id})
# ============================================================

@router.post("/", response_model=HerramientaRead)
def crear_herramienta(herramienta: HerramientaCreate, session: Session = Depends(get_session)):
    db_herr = Herramienta(**herramienta.model_dump())
    session.add(db_herr)
    session.commit()
    session.refresh(db_herr)
    
    # Crear carpeta de la herramienta y escribir .meta.json
    try:
        herr_code = f"H{db_herr.id:04d}"
        nombre_safe = sanitize_filename(db_herr.nombre_herramienta, "SN")
        folder_name = f"{herr_code}_{nombre_safe}"
        herr_dir = get_dir("herramientas_imagenes") / folder_name
        herr_dir.mkdir(parents=True, exist_ok=True)
        meta_data = build_herramienta_meta(db_herr)
        write_meta_json(herr_dir, meta_data)
    except Exception as e:
        print(f"[herramientas.py] WARNING: No se pudo crear .meta.json: {e}")
    
    return db_herr


@router.get("/", response_model=list[HerramientaRead])
def listar_herramientas(session: Session = Depends(get_session)):
    herramientas = session.exec(select(Herramienta)).all()
    return herramientas


# ---------------------------------------------------------
# ENDPOINT: DESCARGAR PLANTILLA EXCEL
# ---------------------------------------------------------
@router.get("/plantilla-excel")
def descargar_plantilla_excel(session: Session = Depends(get_session)):
    """
    Genera y descarga un archivo Excel plantilla con datos de ejemplo
    de herramientas para biolaboratorio.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Herramientas CMMS-BioAI"
    
    encabezados = [
        'nombre_herramienta', 'numero_identificacion', 'descripcion', 'categoria',
        'cantidad_disponible', 'unidad_medida', 'ubicacion_almacen', 'estado_uso',
        'costo_adquisicion', 'fecha_adquisicion', 'proveedor_ultimo', 'observaciones'
    ]
    ws.append(encabezados)
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col_num, _ in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Datos de ejemplo - Herramientas de biolaboratorio
    datos_demo = [
        ["Osciloscopio digital Rigol DS1054Z", "OSC-001", "Osciloscopio 4 canales 50MHz para diagnostico de equipos medicos", "Instrumento de Medición", 1, "unidad", "Taller - Estante A1", "Disponible", 5200.00, "2024-03-15", "Rigol Technologies", "Calibracion anual vigente"],
        ["Multímetro Fluke 87V", "MUL-002", "Multímetro digital True RMS industrial", "Instrumento de Medición", 2, "unidad", "Taller - Estante A1", "Disponible", 2800.00, "2024-01-20", "Fluke Corporation", "Unidad con bateria nueva"],
        ["Analizador de seguridad eléctrica Fluke ESA620", "ANA-003", "Para prueba de seguridad eléctrica en equipos médicos según IEC 62353", "Instrumento de Medición", 1, "unidad", "Taller - Estante A2", "En Uso", 15000.00, "2024-06-10", "Fluke Corporation", "En uso en mantenimiento preventivo"],
        ["Juego de destornilladores aislados Wiha", "DES-004", "Juego 12 piezas VDE 1000V para trabajo en equipos energizados", "Herramienta Manual", 3, "juego", "Taller - Cajón B1", "Disponible", 450.00, "2024-02-05", "Wiha", ""],
        ["Pinza amperimérica Fluke 376 FC", "PIN-005", "Pinza AC/DC 1000A con Bluetooth", "Instrumento de Medición", 1, "unidad", "Taller - Estante A1", "Disponible", 3500.00, "2024-04-18", "Fluke Corporation", "Con certificado de calibracion"],
        ["Alcohol isopropílico 99%", "ALC-006", "Para limpieza de componentes electrónicos y ópticos", "Consumible", 5, "litro", "Taller - Estante C2", "Disponible", 85.00, "2025-01-10", "Quimica Bolivia", "Almacenar en lugar ventilado"],
        ["Estaño para soldadura 60/40 0.8mm", "EST-007", "Rollo 500g soldadura electrónica con núcleo de resina", "Consumible", 4, "rollo", "Taller - Estante C1", "Disponible", 120.00, "2025-02-15", "ElectroSolder", ""],
        ["Kit de calibración presión", "KIT-008", "Kit con manómetro patrón y accesorios para calibración de equipos de presión", "Kit", 1, "kit", "Taller - Estante A3", "Disponible", 8500.00, "2024-08-01", "Additel", "Incluye case y certificados"],
        ["Generador de señales Rigol DG822", "GEN-009", "Generador de funciones 2 canales 30MHz para pruebas de equipos", "Instrumento de Medición", 1, "unidad", "Taller - Estante A2", "Disponible", 4100.00, "2024-05-22", "Rigol Technologies", ""],
        ["Cautín soldador Weller WE1010", "CAU-010", "Estación de soldadura 70W con control de temperatura", "Herramienta Manual", 2, "unidad", "Taller - Estante B1", "Disponible", 680.00, "2024-09-12", "Weller", "Una unidad en reparacion"],
    ]
    
    data_align = Alignment(vertical="center", wrap_text=True)
    
    for fila_data in datos_demo:
        ws.append(fila_data)
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = data_align
    
    anchos = [40, 18, 55, 24, 14, 12, 24, 16, 14, 16, 22, 35]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    
    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA IMPORTAR HERRAMIENTAS"],
        [""],
        ["1. Los campos marcados como obligatorios (*) no pueden estar vacíos:"],
        ["   - nombre_herramienta *", "nombre de la herramienta o material"],
        ["   - cantidad_disponible *", "número entero (ej: 1)"],
        [""],
        ["2. Campo 'categoria' (valores permitidos):"],
        ["   - Instrumento de Medición", "osciloscopios, multímetros, analizadores, etc."],
        ["   - Herramienta Manual", "destornilladores, alicates, soldadores, etc."],
        ["   - Consumible", "alcohol, estaño, cintas, limpiadores, etc."],
        ["   - Kit", "conjuntos de herramientas o accesorios"],
        [""],
        ["3. Campo 'estado_uso' (valores permitidos):"],
        ["   - Disponible", "lista para usar"],
        ["   - En Uso", "actualmente en uso en una OT o actividad"],
        ["   - En Reparación", "fuera de servicio temporalmente"],
        ["   - Dado de Baja", "retirada del servicio permanentemente"],
        [""],
        ["4. Campo 'numero_identificacion': código interno del taller (opcional pero recomendado)."],
        ["5. Campo 'costo_adquisicion': precio en Bs. número decimal (opcional)."],
        ["6. Campo 'fecha_adquisicion': fecha en formato YYYY-MM-DD (opcional)."],
        [""],
        ["7. Límite de archivo: 5MB. Solo archivos .xlsx o .csv."],
        [""],
        ["8. Puede eliminar los datos de ejemplo y usar sus propios datos."],
        ["   Mantenga los encabezados de columna en la fila 1."],
    ]
    for row in instrucciones:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 55
    ws2.column_dimensions['B'].width = 45
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()
    
    filename = f"CMMS-BioAI_Plantilla_Herramientas_{dt.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ---------------------------------------------------------
# ENDPOINT: IMPORTAR HERRAMIENTAS DESDE EXCEL
# ---------------------------------------------------------
@router.post("/import-excel")
async def importar_herramientas_excel(file: UploadFile = File(...), session: Session = Depends(get_session)):
    """
    Importa herramientas desde un archivo Excel (.xlsx) o CSV.
    """
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .csv")
    
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo no debe superar 5MB")
    
    ext = Path(file.filename).suffix.lower() if file.filename else ''
    is_csv = ext == '.csv'
    
    if is_csv:
        try:
            text = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                text = contents.decode('latin-1')
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error al decodificar el archivo CSV: {str(e)}")
        
        try:
            sniffer_sample = text[:2048]
            try:
                dialect = csv.Sniffer().sniff(sniffer_sample, delimiters=',;\t')
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(StringIO(text), dialect)
            filas = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo CSV: {str(e)}")
        
        if len(filas) < 2:
            raise HTTPException(status_code=400, detail="El archivo CSV está vacío o solo tiene encabezados")
    else:
        try:
            wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo Excel: {str(e)}")
    
    COLUMNAS = [
        'nombre_herramienta', 'numero_identificacion', 'descripcion', 'categoria',
        'cantidad_disponible', 'unidad_medida', 'ubicacion_almacen', 'estado_uso',
        'costo_adquisicion', 'fecha_adquisicion', 'proveedor_ultimo', 'observaciones'
    ]
    OBLIGATORIAS = {'nombre_herramienta', 'cantidad_disponible'}
    CATEGORIAS_VALIDAS = {'Instrumento de Medición', 'Herramienta Manual', 'Consumible', 'Kit'}
    ESTADOS_VALIDOS = {'Disponible', 'En Uso', 'En Reparación', 'Dado de Baja'}
    
    if not is_csv:
        ws = None
        for sheet in wb.worksheets:
            filas_tmp = list(sheet.iter_rows(values_only=True))
            if filas_tmp:
                enc_tmp = [str(c).strip().lower() if c else '' for c in filas_tmp[0]]
                if OBLIGATORIAS.issubset(set(enc_tmp)):
                    ws = sheet
                    break
        
        if ws is None:
            ws = wb.worksheets[0]
        
        filas = list(ws.iter_rows(values_only=True))
        if len(filas) < 2:
            raise HTTPException(status_code=400, detail="El archivo está vacío o solo tiene encabezados")
    
    encabezados_leidos = [str(c).strip().lower() if c else '' for c in filas[0]]
    
    col_index = {}
    for i, h in enumerate(encabezados_leidos):
        if h in COLUMNAS:
            col_index[h] = i
    
    faltantes = OBLIGATORIAS - set(col_index.keys())
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas obligatorias: {', '.join(faltantes)}. Encabezados encontrados: {', '.join(encabezados_leidos)}"
        )
    
    exitosos = 0
    actualizados = 0
    fallidos = []
    
    for fila_num, fila in enumerate(filas[1:], start=2):
        try:
            def get_val(col_name):
                idx = col_index.get(col_name)
                if idx is None or idx >= len(fila):
                    return None
                val = fila[idx]
                return str(val).strip() if val is not None else None
            
            errores_fila = []
            for col in OBLIGATORIAS:
                val = get_val(col)
                if not val:
                    errores_fila.append(f"'{col}' es obligatorio")
            
            if errores_fila:
                fallidos.append({
                    "fila": fila_num,
                    "nombre": get_val('nombre_herramienta') or 'N/A',
                    "errores": errores_fila
                })
                continue
            
            # Parsear cantidad_disponible
            try:
                cantidad = int(float(get_val('cantidad_disponible')))
                if cantidad < 0:
                    raise ValueError()
            except (ValueError, TypeError):
                fallidos.append({
                    "fila": fila_num,
                    "nombre": get_val('nombre_herramienta') or 'N/A',
                    "errores": ["cantidad_disponible debe ser un número entero positivo"]
                })
                continue
            
            # Validar categoría
            categoria = get_val('categoria') or 'Herramienta Manual'
            if categoria not in CATEGORIAS_VALIDAS:
                categoria = 'Herramienta Manual'
            
            # Validar estado_uso
            estado_uso = get_val('estado_uso') or 'Disponible'
            if estado_uso not in ESTADOS_VALIDOS:
                estado_uso = 'Disponible'
            
            # Parsear costo_adquisicion
            costo = None
            costo_str = get_val('costo_adquisicion')
            if costo_str:
                try:
                    costo = float(costo_str)
                except (ValueError, TypeError):
                    pass
            
            # Parsear fecha_adquisicion
            fecha_adq = None
            fecha_str = get_val('fecha_adquisicion')
            if fecha_str:
                try:
                    from datetime import date as date_mod
                    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                        try:
                            fecha_adq = date_mod.strptime(str(fecha_str).strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass
            
            unidad = get_val('unidad_medida') or 'unidad'
            
            # Verificar si ya existe (por numero_identificacion o nombre)
            num_id = get_val('numero_identificacion')
            herr_existente = None
            
            if num_id:
                herr_existente = session.exec(
                    select(Herramienta).where(Herramienta.numero_identificacion == num_id)
                ).first()
            
            if not herr_existente:
                nombre = get_val('nombre_herramienta')
                herr_existente = session.exec(
                    select(Herramienta).where(Herramienta.nombre_herramienta == nombre)
                ).first()
            
            if herr_existente:
                # ACTUALIZAR
                herr_existente.nombre_herramienta = get_val('nombre_herramienta') or herr_existente.nombre_herramienta
                herr_existente.numero_identificacion = num_id or herr_existente.numero_identificacion
                herr_existente.descripcion = get_val('descripcion') or herr_existente.descripcion
                herr_existente.categoria = categoria
                herr_existente.cantidad_disponible = cantidad
                herr_existente.unidad_medida = unidad
                herr_existente.ubicacion_almacen = get_val('ubicacion_almacen') or herr_existente.ubicacion_almacen
                herr_existente.estado_uso = estado_uso
                herr_existente.costo_adquisicion = costo if costo is not None else herr_existente.costo_adquisicion
                herr_existente.fecha_adquisicion = fecha_adq or herr_existente.fecha_adquisicion
                herr_existente.proveedor_ultimo = get_val('proveedor_ultimo') or herr_existente.proveedor_ultimo
                herr_existente.observaciones = get_val('observaciones') or herr_existente.observaciones
                session.add(herr_existente)
                actualizados += 1
            else:
                # CREAR nuevo
                nuevo = Herramienta(
                    nombre_herramienta=get_val('nombre_herramienta'),
                    numero_identificacion=num_id,
                    descripcion=get_val('descripcion'),
                    categoria=categoria,
                    cantidad_disponible=cantidad,
                    unidad_medida=unidad,
                    ubicacion_almacen=get_val('ubicacion_almacen'),
                    estado_uso=estado_uso,
                    costo_adquisicion=costo,
                    fecha_adquisicion=fecha_adq,
                    proveedor_ultimo=get_val('proveedor_ultimo'),
                    observaciones=get_val('observaciones')
                )
                session.add(nuevo)
                exitosos += 1
                
        except Exception as e:
            fallidos.append({
                "fila": fila_num,
                "nombre": get_val('nombre_herramienta') if col_index.get('nombre_herramienta') is not None else 'N/A',
                "errores": [str(e)]
            })
    
    session.commit()
    if not is_csv:
        wb.close()
    
    return {
        "exitosos": exitosos,
        "actualizados": actualizados,
        "fallidos": len(fallidos),
        "total_procesados": exitosos + actualizados + len(fallidos),
        "errores": fallidos
    }


# ============================================================
# RUTAS DINÁMICAS (van DESPUÉS de las estáticas)
# ============================================================

# ---------------------------------------------------------
# ENDPOINT: SUBIR IMAGEN PRINCIPAL DE HERRAMIENTA
# Estructura: uploads/HERRAMIENTAS/H0001_nombre/H0001_nombre.ext
# ---------------------------------------------------------
@router.post("/{herr_id}/upload_imagen")
async def upload_imagen_herramienta(herr_id: int, file: UploadFile = File(...), session: Session = Depends(get_session)):
    """Sube una imagen principal para una herramienta."""
    db_herr = session.get(Herramienta, herr_id)
    if not db_herr:
        raise HTTPException(status_code=404, detail="Herramienta no encontrada")

    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp", "image/bmp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Use JPEG, PNG, GIF, WebP o BMP.")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="La imagen no debe superar 5MB")

    herr_code = f"H{herr_id:04d}"
    nombre_safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in (db_herr.nombre_herramienta or "SN"))
    folder_name = f"{herr_code}_{nombre_safe}"

    uploads_dir = get_dir("herramientas_imagenes") / folder_name
    uploads_dir.mkdir(parents=True, exist_ok=True)

    ext = os.path.splitext(file.filename)[1] if file.filename else ".jpg"
    filename = f"{folder_name}{ext}"
    file_path = uploads_dir / filename

    # Eliminar imagen previa con otra extensión
    for old_ext in [".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"]:
        old_path = uploads_dir / f"{folder_name}{old_ext}"
        if old_path.exists() and old_path != file_path:
            old_path.unlink()

    with open(str(file_path), "wb") as f:
        f.write(contents)

    imagen_ruta = f"HERRAMIENTAS/{folder_name}/{filename}"
    db_herr.imagen_ruta = imagen_ruta
    session.add(db_herr)
    session.commit()
    session.refresh(db_herr)

    # Actualizar .meta.json con la nueva imagen
    try:
        meta_data = build_herramienta_meta(db_herr, imagen_ruta)
        write_meta_json(uploads_dir, meta_data)
    except Exception as e:
        print(f"[herramientas.py] WARNING: No se pudo actualizar .meta.json: {e}")

    return {"ok": True, "imagen_ruta": db_herr.imagen_ruta, "nombre_archivo": file.filename}


# ---------------------------------------------------------
# ENDPOINT: ELIMINAR IMAGEN PRINCIPAL DE HERRAMIENTA
# ---------------------------------------------------------
@router.delete("/{herr_id}/imagen")
def eliminar_imagen_herramienta(herr_id: int, session: Session = Depends(get_session)):
    """Elimina la imagen principal de una herramienta (archivo + BD)."""
    db_herr = session.get(Herramienta, herr_id)
    if not db_herr:
        raise HTTPException(status_code=404, detail="Herramienta no encontrada")

    if not db_herr.imagen_ruta:
        raise HTTPException(status_code=400, detail="Esta herramienta no tiene imagen")

    file_path = get_dir("uploads_base") / db_herr.imagen_ruta
    if file_path.exists():
        file_path.unlink()

    db_herr.imagen_ruta = None
    session.add(db_herr)
    session.commit()

    # Actualizar .meta.json (sin imagen)
    try:
        herr_code = f"H{db_herr.id:04d}"
        nombre_safe = sanitize_filename(db_herr.nombre_herramienta, "SN")
        folder_name = f"{herr_code}_{nombre_safe}"
        herr_dir = get_dir("herramientas_imagenes") / folder_name
        if herr_dir.exists():
            meta_data = build_herramienta_meta(db_herr)
            write_meta_json(herr_dir, meta_data)
    except Exception as e:
        print(f"[herramientas.py] WARNING: No se pudo actualizar .meta.json: {e}")

    return {"ok": True, "message": "Imagen eliminada"}


@router.get("/{herr_id}", response_model=HerramientaRead)
def obtener_herramienta(herr_id: int, session: Session = Depends(get_session)):
    db_herr = session.get(Herramienta, herr_id)
    if not db_herr:
        raise HTTPException(status_code=404, detail="Herramienta no encontrada")
    return db_herr


@router.put("/{herr_id}", response_model=HerramientaRead)
def actualizar_herramienta(
    herr_id: int,
    datos: HerramientaUpdate,
    session: Session = Depends(get_session),
):
    db_herr = session.get(Herramienta, herr_id)
    if not db_herr:
        raise HTTPException(status_code=404, detail="Herramienta no encontrada")
    for key, value in datos.model_dump(exclude_unset=True).items():
        setattr(db_herr, key, value)
    session.add(db_herr)
    session.commit()
    session.refresh(db_herr)
    
    # Actualizar .meta.json
    try:
        herr_code = f"H{db_herr.id:04d}"
        nombre_safe = sanitize_filename(db_herr.nombre_herramienta, "SN")
        folder_name = f"{herr_code}_{nombre_safe}"
        herr_dir = get_dir("herramientas_imagenes") / folder_name
        if herr_dir.exists():
            meta_data = build_herramienta_meta(db_herr, db_herr.imagen_ruta)
            write_meta_json(herr_dir, meta_data)
    except Exception as e:
        print(f"[herramientas.py] WARNING: No se pudo actualizar .meta.json: {e}")
    
    return db_herr


@router.delete("/{herr_id}")
def eliminar_herramienta(herr_id: int, session: Session = Depends(get_session)):
    db_herr = session.get(Herramienta, herr_id)
    if not db_herr:
        raise HTTPException(status_code=404, detail="Herramienta no encontrada")
    session.delete(db_herr)
    session.commit()
    return {"ok": True, "message": "Herramienta eliminada"}
