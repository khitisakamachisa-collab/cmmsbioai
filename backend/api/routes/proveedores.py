from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from database import get_session
from models.proveedores import Proveedor, ContactoProveedor
from schemas.proveedor import (
    ProveedorCreate, ProveedorUpdate, ProveedorRead, ProveedorReadWithContactos,
    ContactoProveedorCreate, ContactoProveedorUpdate, ContactoProveedorRead,
)
from io import BytesIO, StringIO
from datetime import datetime as dt
from pathlib import Path
import openpyxl
import csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/proveedores", tags=["Proveedores"])


# ============================================================
# RUTAS ESTÁTICAS (deben ir ANTES de /{proveedor_id})
# ============================================================

# ---------------------------------------------------------
# ENDPOINT: DESCARGAR PLANTILLA CSV
# ---------------------------------------------------------
@router.get("/plantilla-csv")
def descargar_plantilla_csv(session: Session = Depends(get_session)):
    """
    [RESPALDO] La fuente principal de esta plantilla ahora es el archivo estatico en:
        frontend/public/plantillas/plantilla_proveedores.csv
    El frontend descarga directamente desde ahi (sin llamar al backend).
    Este endpoint se mantiene como respaldo y para documentacion Swagger.

    Genera y descarga un archivo CSV plantilla con datos de ejemplo
    de proveedores biomédicos para Bolivia.
    """
    output = StringIO()
    encabezados = [
        'nombre_empresa', 'ciudad', 'direccion',
        'telefono_principal', 'email_principal', 'pagina_web', 'notas_generales'
    ]

    writer = csv.writer(output)
    writer.writerow(encabezados)

    datos_demo = [
        ["TechMed Bolivia SRL", "Cochabamba", "Av. Blanco Galindo km 7.5", "+591 4 4223344", "ventas@techmed.bo", "https://techmed.bo", "Distribuidor autorizado de equipos de monitoreo"],
        ["BioSupply Andina", "La Paz", "Calle Mercado N° 1234, Zona Sopocachi", "+591 2 2445566", "info@biosupply.bo", "https://biosupply.bo", "Repuestos y consumibles para laboratorio clínico"],
        ["MedEquip Bolivia SA", "Santa Cruz", "Tercer Anillo Interno, Av. Roca y Coronado", "+591 3 3467788", "ventas@medequip.bo", "", "Importador de equipos de imagenología"],
        ["LabTech SRL", "Cochabamba", "Av. América Esq. Esteban Arze", "+591 4 4521199", "contacto@labtech.bo", "https://labtech.bo", "Especialistas en equipos de laboratorio"],
    ]

    for fila in datos_demo:
        writer.writerow(fila)

    output.seek(0)
    filename = f"CMMS-BioAI_Plantilla_Proveedores_{dt.now().strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "text/csv; charset=utf-8"
        }
    )


# ---------------------------------------------------------
# ENDPOINT: DESCARGAR PLANTILLA EXCEL DEMO
# ---------------------------------------------------------
@router.get("/plantilla-excel")
def descargar_plantilla_excel(session: Session = Depends(get_session)):
    """
    [RESPALDO] La fuente principal de esta plantilla ahora es el archivo estatico en:
        frontend/public/plantillas/plantilla_proveedores.xlsx
    El frontend descarga directamente desde ahi (sin llamar al backend).
    Este endpoint se mantiene como respaldo y para documentacion Swagger.

    Genera y descarga un archivo Excel plantilla con datos de ejemplo
    de proveedores biomédicos en Bolivia.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores CMMS-BioAI"

    encabezados = [
        'nombre_empresa', 'ciudad', 'direccion',
        'telefono_principal', 'email_principal', 'pagina_web', 'notas_generales'
    ]
    ws.append(encabezados)

    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, _ in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Datos de ejemplo - Proveedores del rubro biomédico en Bolivia
    datos_demo = [
        ["TechMed Bolivia SRL", "Cochabamba", "Av. Blanco Galindo km 7.5", "+591 4 4223344", "ventas@techmed.bo", "https://techmed.bo", "Distribuidor autorizado de equipos de monitoreo y signos vitales"],
        ["BioSupply Andina", "La Paz", "Calle Mercado N° 1234, Zona Sopocachi", "+591 2 2445566", "info@biosupply.bo", "https://biosupply.bo", "Repuestos y consumibles para laboratorio clínico"],
        ["MedEquip Bolivia SA", "Santa Cruz", "Tercer Anillo Interno, Av. Roca y Coronado", "+591 3 3467788", "ventas@medequip.bo", "", "Importador de equipos de imagenología (ecógrafos, rayos X)"],
        ["LabTech SRL", "Cochabamba", "Av. América Esq. Esteban Arze", "+591 4 4521199", "contacto@labtech.bo", "https://labtech.bo", "Especialistas en equipos de laboratorio: centrífugas, espectrofotómetros"],
        ["Hospimed Bolivia", "La Paz", "Av. 6 de Agosto N° 2115", "+591 2 2150500", "ventas@hospimed.bo", "https://hospimed.bo", "Equipos de terapia intensiva y monitoreo multiparámetro"],
        ["DentalPro Suministros", "Cochabamba", "Calle Juan Caprario N° 0456", "+591 4 4256677", "info@dentalpro.bo", "https://dentalpro.bo", "Unidades dentales, autoclaves y consumibles"],
        ["Roche Diagnostics Bolivia", "La Paz", "Av. Mariscal Santa Cruz, Edif. Cámara Nacional", "+591 2 2334455", "bolivia@roche.com", "https://diagnostics.roche.com", "Reactivos y equipos de diagnóstico in vitro"],
        ["Thermo Fisher Andina", "Santa Cruz", "Parque Industrial Mz 7 Lote 12", "+591 3 3528899", "andina@thermofisher.com", "https://thermofisher.com", "Incubadoras, refrigeradores de laboratorio, centrífugas"],
        ["Steris Bolivia SA", "Cochabamba", "Av. Petrolera km 4.5", "+591 4 4710022", "servicio@steris.bo", "https://steris.bo", "Esterilizadores y autoclaves hospitalarias"],
        ["Olympus Latin America", "La Paz", "Av. Arce N° 2798, Edif. Multicentro", "+591 2 2773344", "latam@olympus.com", "https://olympus-latinamerica.com", "Microscopios, endoscopios y equipamiento óptico médico"],
        ["Mindray Bolivia", "Santa Cruz", "Av. San Martín N° 1700", "+591 3 3391212", "bolivia@mindray.com", "https://mindray.com", "Monitores de pacientes, ventiladores y equipos de ultrasonido"],
        ["GE Healthcare Bolivia", "Cochabamba", "Av. Oquendo N° 1245", "+591 4 4533344", "bolivia@gehealthcare.com", "https://gehealthcare.com", "Equipos de imágenes médicas: TAC, MRI, rayos X digitales"],
        ["Zoll Medical Andina", "La Paz", "Calle Colón N° 1590", "+591 2 2314567", "andina@zoll.com", "https://zoll.com", "Desfibriladores, monitores de reanimación y equipos de emergencia"],
        ["Hamilton Bolivia", "Cochabamba", "Av. Blanco Galindo N° 8890", "+591 4 4719988", "bolivia@hamiltoncompany.com", "https://hamiltoncompany.com", "Pipetas, dispensadores y consumibles de precisión"],
        ["Eppendorf Andina", "Santa Cruz", "Av. Monseñor Rivero N° 550", "+591 3 3332211", "andina@eppendorf.com", "https://eppendorf.com", "Microcentrífugas, pipetas, tubos y consumibles de laboratorio"],
        ["Hanna Instruments Bolivia", "La Paz", "Av. 16 de Julio N° 1730", "+591 2 2288776", "bolivia@hannainst.com", "https://hannainst.com", "Medidores de pH, conductividad y oxígeno disuelto"],
        ["3M Healthcare Bolivia", "Cochabamba", "Av. Cristo Redentor N° 0990", "+591 4 4785566", "bolivia@3m.com", "https://3m.com", "Electrodos ECG, productos de esterilización y materiales médicos"],
        ["Vidriolab SRL", "Cochabamba", "Calle Sucre N° 0050", "+591 4 4513344", "ventas@vidriolab.bo", "", "Vidriería de laboratorio: tubos, matraces, pipetas de vidrio"],
        ["Sysmex Latin America", "La Paz", "Av. Ballivián N° 700", "+591 2 2126677", "latam@sysmex.com", "https://sysmex.com", "Analizadores hematológicos y reactivos"],
        ["PaperMed Bolivia", "Santa Cruz", "Av. Argentida N° 2320", "+591 3 3367788", "info@papermed.bo", "", "Papel térmico para monitores, formularios continuos y consumibles"],
    ]

    data_align = Alignment(vertical="center", wrap_text=True)

    for fila_data in datos_demo:
        ws.append(fila_data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = data_align

    anchos = [28, 14, 45, 16, 28, 28, 60]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA IMPORTAR PROVEEDORES"],
        [""],
        ["1. Campo obligatorio (*) - no puede estar vacío:"],
        ["   - nombre_empresa *", "nombre legal o comercial del proveedor (debe ser único)"],
        [""],
        ["2. Campo 'ciudad': ciudad principal donde opera el proveedor (ej: La Paz, Cochabamba, Santa Cruz)."],
        ["   Es independiente de 'direccion' y se usa para filtros."],
        [""],
        ["3. Campo 'direccion': dirección física completa de la empresa (calle, número, zona, etc.)."],
        ["4. Campo 'telefono_principal': teléfono de la empresa (con código de país +591 si aplica)."],
        ["5. Campo 'email_principal': correo electrónico principal de la empresa."],
        ["6. Campo 'pagina_web': URL del sitio web oficial (opcional)."],
        ["7. Campo 'notas_generales': observaciones comerciales, garantías, condiciones, etc."],
        [""],
        ["8. Si el 'nombre_empresa' ya existe en la base de datos, se ACTUALIZARÁ el registro (upsert)."],
        ["   Si no existe, se creará un nuevo proveedor."],
        [""],
        ["9. Los CONTACTOS asociados no se importan desde este Excel. Una vez importados los"],
        ["   proveedores, agrega contactos individualmente desde la página de Proveedores."],
        [""],
        ["10. Límite de archivo: 5MB. Solo archivos .xlsx o .csv."],
        [""],
        ["11. Puede eliminar los datos de ejemplo y usar sus propios datos."],
        ["    Mantenga los encabezados de columna en la fila 1."],
    ]
    for row in instrucciones:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 65
    ws2.column_dimensions['B'].width = 55

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()

    filename = f"CMMS-BioAI_Plantilla_Proveedores_{dt.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ---------------------------------------------------------
# ENDPOINT: IMPORTAR PROVEEDORES DESDE EXCEL
# ---------------------------------------------------------
@router.post("/import-excel")
async def importar_proveedores_excel(file: UploadFile = File(...), session: Session = Depends(get_session)):
    """
    Importa proveedores desde un archivo Excel (.xlsx) o CSV.
    Columnas esperadas (encabezado en fila 1):
      nombre_empresa *, ciudad, direccion,
      telefono_principal, email_principal, pagina_web, notas_generales

    - Campo obligatorio: nombre_empresa
    - Si nombre_empresa ya existe, se ACTUALIZA el registro (upsert)
    """
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .csv")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo no debe superar 5MB")

    ext = Path(file.filename).suffix.lower() if file.filename else ''
    is_csv = ext == '.csv'

    if is_csv:
        try:
            text = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                text = contents.decode('latin-1')
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error al decodificar el archivo CSV: {str(e)}")

        try:
            sniffer_sample = text[:2048]
            try:
                dialect = csv.Sniffer().sniff(sniffer_sample, delimiters=',;\t')
            except csv.Error:
                dialect = csv.excel

            reader = csv.reader(StringIO(text), dialect)
            filas = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo CSV: {str(e)}")

        if len(filas) < 2:
            raise HTTPException(status_code=400, detail="El archivo CSV está vacío o solo tiene encabezados")
    else:
        try:
            wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo Excel: {str(e)}")

    COLUMNAS = [
        'nombre_empresa', 'ciudad', 'direccion',
        'telefono_principal', 'email_principal', 'pagina_web', 'notas_generales'
    ]
    OBLIGATORIAS = {'nombre_empresa'}

    # Buscar la hoja con encabezados (solo Excel)
    if not is_csv:
        ws = None
        for sheet in wb.worksheets:
            filas_tmp = list(sheet.iter_rows(values_only=True))
            if filas_tmp:
                enc_tmp = [str(c).strip().lower() if c else '' for c in filas_tmp[0]]
                if OBLIGATORIAS.issubset(set(enc_tmp)):
                    ws = sheet
                    break

        if ws is None:
            ws = wb.worksheets[0]

        filas = list(ws.iter_rows(values_only=True))
        if len(filas) < 2:
            raise HTTPException(status_code=400, detail="El archivo está vacío o solo tiene encabezados")

    encabezados_leidos = [str(c).strip().lower() if c else '' for c in filas[0]]

    col_index = {}
    for i, h in enumerate(encabezados_leidos):
        if h in COLUMNAS:
            col_index[h] = i

    faltantes = OBLIGATORIAS - set(col_index.keys())
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas obligatorias: {', '.join(faltantes)}. Encabezados encontrados: {', '.join(encabezados_leidos)}"
        )

    exitosos = 0
    actualizados = 0
    fallidos = []

    for fila_num, fila in enumerate(filas[1:], start=2):
        try:
            def get_val(col_name):
                idx = col_index.get(col_name)
                if idx is None or idx >= len(fila):
                    return None
                val = fila[idx]
                return str(val).strip() if val is not None else None

            errores_fila = []
            for col in OBLIGATORIAS:
                val = get_val(col)
                if not val:
                    errores_fila.append(f"'{col}' es obligatorio")

            if errores_fila:
                fallidos.append({
                    "fila": fila_num,
                    "nombre": get_val('nombre_empresa') or 'N/A',
                    "errores": errores_fila
                })
                continue

            nombre = get_val('nombre_empresa')

            # Buscar proveedor existente por nombre_empresa (upsert)
            prov_existente = session.exec(
                select(Proveedor).where(Proveedor.nombre_empresa == nombre)
            ).first()

            if prov_existente:
                # ACTUALIZAR
                prov_existente.ciudad = get_val('ciudad') or prov_existente.ciudad
                prov_existente.direccion = get_val('direccion') or prov_existente.direccion
                prov_existente.telefono_principal = get_val('telefono_principal') or prov_existente.telefono_principal
                prov_existente.email_principal = get_val('email_principal') or prov_existente.email_principal
                prov_existente.pagina_web = get_val('pagina_web') or prov_existente.pagina_web
                prov_existente.notas_generales = get_val('notas_generales') or prov_existente.notas_generales
                session.add(prov_existente)
                actualizados += 1
            else:
                # CREAR nuevo
                nuevo = Proveedor(
                    nombre_empresa=nombre,
                    ciudad=get_val('ciudad'),
                    direccion=get_val('direccion'),
                    telefono_principal=get_val('telefono_principal'),
                    email_principal=get_val('email_principal'),
                    pagina_web=get_val('pagina_web'),
                    notas_generales=get_val('notas_generales'),
                )
                session.add(nuevo)
                exitosos += 1

        except Exception as e:
            fallidos.append({
                "fila": fila_num,
                "nombre": get_val('nombre_empresa') if col_index.get('nombre_empresa') is not None else 'N/A',
                "errores": [str(e)]
            })

    session.commit()
    if not is_csv:
        wb.close()

    return {
        "exitosos": exitosos,
        "actualizados": actualizados,
        "fallidos": len(fallidos),
        "total_procesados": exitosos + actualizados + len(fallidos),
        "errores": fallidos
    }


# ============================================================
# PROVEEDORES - CRUD principal
# ============================================================
@router.post("/", response_model=ProveedorRead, status_code=201)
def crear_proveedor(payload: ProveedorCreate, session: Session = Depends(get_session)):
    """Crea un nuevo proveedor."""
    existe = session.exec(
        select(Proveedor).where(Proveedor.nombre_empresa == payload.nombre_empresa)
    ).first()
    if existe:
        raise HTTPException(status_code=400, detail="Ya existe un proveedor con ese nombre de empresa")

    db_proveedor = Proveedor(**payload.model_dump())
    session.add(db_proveedor)
    session.commit()
    session.refresh(db_proveedor)
    return db_proveedor


@router.get("/", response_model=list[ProveedorRead])
def listar_proveedores(session: Session = Depends(get_session)):
    """Lista todos los proveedores (sin contactos)."""
    return session.exec(select(Proveedor).order_by(Proveedor.nombre_empresa)).all()


@router.get("/con-contactos", response_model=list[ProveedorReadWithContactos])
def listar_proveedores_con_contactos(session: Session = Depends(get_session)):
    """Lista todos los proveedores incluyendo sus contactos asociados."""
    proveedores = session.exec(select(Proveedor).order_by(Proveedor.nombre_empresa)).all()
    resultado = []
    for prov in proveedores:
        contactos = session.exec(
            select(ContactoProveedor).where(ContactoProveedor.proveedor_id == prov.id)
        ).all()
        prov_dict = ProveedorReadWithContactos.model_validate(prov)
        prov_dict.contactos = contactos
        resultado.append(prov_dict)
    return resultado


@router.get("/{proveedor_id}", response_model=ProveedorReadWithContactos)
def obtener_proveedor(proveedor_id: int, session: Session = Depends(get_session)):
    """Obtiene un proveedor por ID incluyendo sus contactos."""
    proveedor = session.get(Proveedor, proveedor_id)
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    contactos = session.exec(
        select(ContactoProveedor).where(ContactoProveedor.proveedor_id == proveedor_id)
    ).all()
    prov_dict = ProveedorReadWithContactos.model_validate(proveedor)
    prov_dict.contactos = contactos
    return prov_dict


@router.put("/{proveedor_id}", response_model=ProveedorRead)
def actualizar_proveedor(proveedor_id: int, payload: ProveedorUpdate, session: Session = Depends(get_session)):
    """Actualiza los datos de un proveedor."""
    db_proveedor = session.get(Proveedor, proveedor_id)
    if not db_proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    datos = payload.model_dump(exclude_unset=True)
    # Validar unicidad del nombre si se esta cambiando
    if "nombre_empresa" in datos and datos["nombre_empresa"] != db_proveedor.nombre_empresa:
        existe = session.exec(
            select(Proveedor).where(Proveedor.nombre_empresa == datos["nombre_empresa"])
        ).first()
        if existe:
            raise HTTPException(status_code=400, detail="Ya existe un proveedor con ese nombre de empresa")

    for key, value in datos.items():
        setattr(db_proveedor, key, value)

    session.add(db_proveedor)
    session.commit()
    session.refresh(db_proveedor)
    return db_proveedor


@router.delete("/{proveedor_id}", status_code=204)
def eliminar_proveedor(proveedor_id: int, session: Session = Depends(get_session)):
    """Elimina un proveedor y todos sus contactos asociados."""
    db_proveedor = session.get(Proveedor, proveedor_id)
    if not db_proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    # Eliminar contactos asociados primero (cascade manual)
    contactos = session.exec(
        select(ContactoProveedor).where(ContactoProveedor.proveedor_id == proveedor_id)
    ).all()
    for c in contactos:
        session.delete(c)

    # v0.9.11: Desasociar equipos cuyo proveedor_principal era este proveedor
    # (no eliminamos los equipos, solo quitamos la referencia)
    from models.equipos import Equipo
    equipos_asociados = session.exec(
        select(Equipo).where(Equipo.proveedor_principal_id == proveedor_id)
    ).all()
    for eq in equipos_asociados:
        eq.proveedor_principal_id = None
        session.add(eq)

    session.delete(db_proveedor)
    session.commit()
    return None


# ============================================================
# v0.9.11: EQUIPOS ASOCIADOS AL PROVEEDOR
# Endpoint para asociar/desasociar equipos en lote usando
# el campo Equipo.proveedor_principal_id (relación 1:N)
# ============================================================
@router.get("/{proveedor_id}/equipos")
def listar_equipos_proveedor(proveedor_id: int, session: Session = Depends(get_session)):
    """
    Lista los equipos cuyo proveedor_principal_id es el proveedor indicado.
    Útil para mostrar los equipos asociados a un proveedor en el frontend.
    """
    from models.equipos import Equipo
    proveedor = session.get(Proveedor, proveedor_id)
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    equipos = session.exec(
        select(Equipo).where(Equipo.proveedor_principal_id == proveedor_id)
    ).all()
    return [
        {
            "id": eq.id,
            "nombre_corto": eq.nombre_corto,
            "modelo": eq.modelo,
            "marca": eq.marca,
            "numero_serie": eq.numero_serie,
            "ubicacion_actual": eq.ubicacion_actual,
        }
        for eq in equipos
    ]


@router.put("/{proveedor_id}/equipos")
def actualizar_equipos_proveedor(
    proveedor_id: int,
    payload: dict,
    session: Session = Depends(get_session)
):
    """
    Sincroniza los equipos cuyo proveedor_principal es este proveedor.
    Recibe: { "equipos_ids": [1, 2, 3] }

    Lógica:
    - Equipos en la lista que tenían OTRO proveedor → se asignan a este.
    - Equipos que estaban asignados a este pero NO están en la lista → se desasocian (NULL).
    - Equipos en la lista que YA estaban asignados a este → no se tocan.
    """
    from models.equipos import Equipo
    proveedor = session.get(Proveedor, proveedor_id)
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    equipos_ids_nuevos = payload.get("equipos_ids") or []
    if not isinstance(equipos_ids_nuevos, list):
        raise HTTPException(status_code=400, detail="equipos_ids debe ser una lista de enteros")

    # Convertir a set de ints
    try:
        nuevos_set = {int(x) for x in equipos_ids_nuevos}
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="equipos_ids debe contener enteros válidos")

    # Equipos actualmente asignados a este proveedor
    actuales = session.exec(
        select(Equipo).where(Equipo.proveedor_principal_id == proveedor_id)
    ).all()
    actuales_ids = {eq.id for eq in actuales}

    # Desasociar los que ya no están en la nueva lista
    for eq in actuales:
        if eq.id not in nuevos_set:
            eq.proveedor_principal_id = None
            session.add(eq)

    # Asociar los nuevos (verificando que existan)
    for eq_id in nuevos_set:
        eq = session.get(Equipo, eq_id)
        if eq:
            eq.proveedor_principal_id = proveedor_id
            session.add(eq)

    session.commit()

    return {
        "ok": True,
        "proveedor_id": proveedor_id,
        "equipos_asociados": len(nuevos_set),
        "mensaje": f"Se asociaron {len(nuevos_set)} equipo(s) al proveedor"
    }


# ============================================================
# CONTACTOS DEL PROVEEDOR - Sub-recurso anidado
# ============================================================
@router.post("/{proveedor_id}/contactos", response_model=ContactoProveedorRead, status_code=201)
def agregar_contacto(proveedor_id: int, payload: ContactoProveedorCreate, session: Session = Depends(get_session)):
    """Agrega un contacto a un proveedor existente."""
    proveedor = session.get(Proveedor, proveedor_id)
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")

    datos = payload.model_dump(exclude_unset=True)
    datos["proveedor_id"] = proveedor_id
    db_contacto = ContactoProveedor(**datos)
    session.add(db_contacto)
    session.commit()
    session.refresh(db_contacto)
    return db_contacto


@router.get("/{proveedor_id}/contactos", response_model=list[ContactoProveedorRead])
def listar_contactos_de_proveedor(proveedor_id: int, session: Session = Depends(get_session)):
    """Lista todos los contactos asociados a un proveedor."""
    proveedor = session.get(Proveedor, proveedor_id)
    if not proveedor:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    return session.exec(
        select(ContactoProveedor).where(ContactoProveedor.proveedor_id == proveedor_id)
    ).all()


@router.put("/contactos/{contacto_id}", response_model=ContactoProveedorRead)
def actualizar_contacto(contacto_id: int, payload: ContactoProveedorUpdate, session: Session = Depends(get_session)):
    """Actualiza un contacto especifico."""
    db_contacto = session.get(ContactoProveedor, contacto_id)
    if not db_contacto:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")

    datos = payload.model_dump(exclude_unset=True)
    for key, value in datos.items():
        setattr(db_contacto, key, value)

    session.add(db_contacto)
    session.commit()
    session.refresh(db_contacto)
    return db_contacto


@router.delete("/contactos/{contacto_id}", status_code=204)
def eliminar_contacto(contacto_id: int, session: Session = Depends(get_session)):
    """Elimina un contacto especifico."""
    db_contacto = session.get(ContactoProveedor, contacto_id)
    if not db_contacto:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")

    session.delete(db_contacto)
    session.commit()
    return None
