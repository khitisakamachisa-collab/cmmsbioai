from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from database import get_session
from models.repuestos import Repuesto
from schemas.repuesto import RepuestoCreate, RepuestoRead, RepuestoUpdate
from io import BytesIO, StringIO
from datetime import datetime as dt
from pathlib import Path
import openpyxl
import csv
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import get_dir

router = APIRouter(prefix="/repuestos", tags=["Inventario"])

# ============================================================
# RUTAS ESTÁTICAS (deben ir ANTES de /{rep_id})
# ============================================================

@router.post("/", response_model=RepuestoRead)
def crear_repuesto(repuesto: RepuestoCreate, session: Session = Depends(get_session)):
    db_rep = Repuesto(**repuesto.model_dump())
    session.add(db_rep)
    session.commit()
    session.refresh(db_rep)
    return db_rep

@router.get("/", response_model=list[RepuestoRead])
def listar_repuestos(session: Session = Depends(get_session)):
    repuestos = session.exec(select(Repuesto)).all()
    return repuestos


# ---------------------------------------------------------
# ENDPOINT: DESCARGAR PLANTILLA CSV
# ---------------------------------------------------------
@router.get("/plantilla-csv")
def descargar_plantilla_csv(session: Session = Depends(get_session)):
    """
    Genera y descarga un archivo CSV plantilla con datos de ejemplo
    de repuestos para biolaboratorio.
    """
    output = StringIO()
    encabezados = [
        'nombre_repuesto', 'numero_serie', 'numero_material',
        'descripcion', 'especificaciones_tecnicas',
        'cantidad_disponible', 'unidad_medida', 'ubicacion_almacen', 'nivel_stock_minimo',
        'proveedor_ultimo', 'fecha_ultima_entrada', 'precio_referencia'
    ]
    
    writer = csv.writer(output)
    writer.writerow(encabezados)
    
    datos_demo = [
        ["Filtro HEPA para cabina de flujo", "SN-FLT-001", "FLT-HEP-001", "Filtro HEPA 99.97% 0.3um para cabina Esco A2", "99.97% 0.3um, tamano estandar", 3, "unidad", "Almacen B - Estante 1", 2, "BioSupply SRL", "2025-01-15", 850.50],
        ["Lampara UV para autoclave", "SN-LMP-002", "LMP-UV-002", "Lampara germicida UV 254nm Steris", "254nm, 15W, T8", 5, "unidad", "Almacen B - Estante 1", 2, "MedEquip Bolivia", "2025-02-20", 320.00],
        ["Electrodo pH recalculado", "SN-ELC-003", "ELC-PH-003", "Electrodo de pH combinado para analizadores", "Combinado, rango 0-14 pH", 4, "unidad", "Lab. Bioquimica", 2, "LabTech SA", "2025-03-10", 1200.00],
    ]
    
    for fila in datos_demo:
        writer.writerow(fila)
    
    output.seek(0)
    filename = f"CMMS-BioAI_Plantilla_Repuestos_{dt.now().strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "text/csv; charset=utf-8"
        }
    )


# ---------------------------------------------------------
# ENDPOINT: DESCARGAR PLANTILLA EXCEL DEMO
# ---------------------------------------------------------
@router.get("/plantilla-excel")
def descargar_plantilla_excel(session: Session = Depends(get_session)):
    """
    Genera y descarga un archivo Excel plantilla con datos de ejemplo
    de repuestos para biolaboratorio.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Repuestos CMMS-BioAI"
    
    encabezados = [
        'nombre_repuesto', 'numero_serie', 'numero_material',
        'descripcion', 'especificaciones_tecnicas',
        'cantidad_disponible', 'unidad_medida', 'ubicacion_almacen', 'nivel_stock_minimo',
        'proveedor_ultimo', 'fecha_ultima_entrada', 'precio_referencia'
    ]
    ws.append(encabezados)
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col_num, _ in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Datos de ejemplo - Repuestos de biolaboratorio
    datos_demo = [
        ["Filtro HEPA para cabina de flujo", "SN-FLT-001", "FLT-HEP-001", "Filtro HEPA 99.97% 0.3μm para cabina Esco A2", "99.97% 0.3um, tamano estandar", 3, "unidad", "Almacén B - Estante 1", 2, "BioSupply SRL", "2025-01-15", 850.50],
        ["Lámpara UV para autoclave", "SN-LMP-002", "LMP-UV-002", "Lámpara germicida UV 254nm Steris", "254nm, 15W, T8", 5, "unidad", "Almacén B - Estante 1", 2, "MedEquip Bolivia", "2025-02-20", 320.00],
        ["Electrodo pH recalculado", "SN-ELC-003", "ELC-PH-003", "Electrodo de pH combinado para analizadores", "Combinado, rango 0-14 pH", 4, "unidad", "Lab. Bioquímica", 2, "LabTech SA", "2025-03-10", 1200.00],
        ["Tubos de ensayo 16x100mm", "", "TUB-16004", "Tubos de vidrio borosilicato, paquete x100", "16x100mm, borosilicato", 25, "paquete", "Almacén A - Estante 3", 10, "Vidriolab", "2025-04-01", 45.00],
        ["Reactivos panel bioquímico", "", "RCT-BIO-005", "Kit de reactivos Cobas c311 panel completo", "Panel completo, vigencia 6 meses", 2, "kit", "Refrigerador Lab 2°C-8°C", 1, "Roche Diagnostics", "2025-05-15", 3500.00],
        ["Filtro de aire compresor", "", "FLT-AIR-006", "Filtro de línea para compresor dental/medico", "5 micras, roscable 1/4", 6, "unidad", "Almacén B - Estante 2", 3, "DentalPro", "2025-03-22", 120.00],
        ["Cable ECG 12 derivaciones", "SN-CBL-007", "CBL-ECG-007", "Cable paciente 12-lead para GE MAC 2000", "12-lead, conector GE", 3, "unidad", "Almacén C - Gabinete 1", 2, "GE Healthcare", "2025-01-30", 2100.00],
        ["Sensor SpO2 dedo", "", "SNR-SPO-008", "Sensor pulsioximetría dedo adulto Mindray", "Dedo adulto, compatible Mindray", 4, "unidad", "UCI - Armario suministros", 2, "Mindray", "2025-04-10", 780.00],
        ["Papel térmico monitor", "", "PPL-TRM-009", "Papel térmico 112mm para monitor de signos vitales", "112mm x 30m", 8, "rollo", "Almacén A - Estante 5", 3, "PaperMed", "2025-06-01", 25.00],
        ["Batería interna desfibrilador", "SN-BAT-010", "BAT-DEF-010", "Batería recargable Zoll R Series 14.4V", "14.4V Li-Ion, autonomia 8h", 2, "unidad", "Emergencias - Armario", 1, "Zoll Medical", "2024-12-15", 4500.00],
        ["Goma/pistón pipeta", "", "GOM-PIP-011", "Pistón de goma para pipeta Hamilton 100-1000μL", "Compatible Hamilton 100-1000uL", 10, "unidad", "Lab. Automatización", 5, "Hamilton", "2025-05-20", 35.00],
        ["Sellador de tapa incubadora", "", "SLR-INC-012", "Junta de goma puerta incubadora Thermo Heracell", "Silicona, tamano estandar Heracell", 2, "unidad", "Almacén B - Estante 4", 1, "Thermo Fisher", "2025-02-28", 290.00],
        ["Fluorescente microscopio", "", "LMP-MIC-013", "Lámpara halógena 6V 30W para Olympus CX23", "6V 30W, base BA15d", 3, "unidad", "Almacén B - Estante 1", 2, "Olympus", "2025-03-15", 180.00],
        ["Reactivos hematología", "", "RCT-HEM-014", "Kit diluyente/limpiador analizador hematológico", "Diluyente + limpiador + hemolizante", 3, "kit", "Refrigerador Lab 2°C-8°C", 1, "Sysmex", "2025-04-20", 2800.00],
        ["Filtro CO2 incubadora", "", "FLT-CO2-015", "Filtro HEPA para incubadora CO2 Thermo", "HEPA, 99.97% para Thermo 150/250", 2, "unidad", "Almacén B - Estante 3", 1, "Thermo Fisher", "2025-01-25", 650.00],
        ["Aceite bomba vacío", "", "ACE-VAC-016", "Aceite para bomba de vacío Autoclave Steris 1L", "1L, grado vacío", 4, "litro", "Almacén A - Estante 2", 2, "Steris", "2025-05-05", 150.00],
        ["Manguera silicona", "", "MNG-SIL-017", "Tubo silicona 8mm ID para bombas de infusión", "8mm ID, medical grade", 5, "metro", "Almacén C - Estante 1", 3, "MedSilicone", "2025-04-15", 18.00],
        ["Electrodo ECG desechable", "", "ELC-ECG-018", "Electrodos Ag/Cl adulto paquete x50", "Ag/Cl, adulto, x50", 20, "paquete", "UCI - Armario suministros", 10, "3M Healthcare", "2025-06-10", 55.00],
        ["Papel filtro centrífuga", "", "PPL-CEN-019", "Filtro de carbón para Eppendorf 5424", "Carbón activado, compatible 5424", 6, "unidad", "Lab. Hematología", 3, "Eppendorf", "2025-03-30", 95.00],
        ["Solución calibración pH", "", "SLC-PH-020", "Solución buffer pH 4.01 / 7.00 / 10.01 set", "Set 3 botellas 500mL", 5, "kit", "Lab. Control Calidad", 2, "Hanna Instruments", "2025-02-10", 280.00],
    ]
    
    data_align = Alignment(vertical="center", wrap_text=True)
    
    for fila_data in datos_demo:
        ws.append(fila_data)
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = data_align
    
    anchos = [35, 16, 18, 50, 30, 14, 12, 28, 18, 22, 18, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    
    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA IMPORTAR REPUESTOS"],
        [""],
        ["1. Los campos marcados como obligatorios (*) no pueden estar vacíos:"],
        ["   - nombre_repuesto *", "nombre del repuesto o insumo"],
        ["   - cantidad_disponible *", "número entero (ej: 5)"],
        [""],
        ["2. Campo 'numero_material': código del repuesto. Si ya existe, se ACTUALIZARÁ."],
        ["   Si no existe numero_material, se busca por nombre_repuesto."],
        [""],
        ["3. Campo 'numero_serie': número de serie del repuesto (opcional)."],
        ["4. Campo 'especificaciones_tecnicas': voltaje, tamaño, capacidad, etc. (opcional)."],
        ["5. Campo 'nivel_stock_minimo': entero. Cantidad mínima antes de alerta."],
        ["   Si no se especifica, no se generará alerta de stock bajo."],
        [""],
        ["6. Campo 'unidad_medida': unidad, par, metro, litro, kit, paquete, rollo, etc."],
        ["   Si se deja vacío, se usará 'unidad' por defecto."],
        [""],
        ["7. Campo 'proveedor_ultimo': nombre del último proveedor (opcional)."],
        ["8. Campo 'fecha_ultima_entrada': fecha en formato YYYY-MM-DD (opcional)."],
        ["9. Campo 'precio_referencia': precio en Bs. número decimal (opcional)."],
        [""],
        ["10. Límite de archivo: 5MB. Solo archivos .xlsx o .csv."],
        [""],
        ["11. Puede eliminar los datos de ejemplo y usar sus propios datos."],
        ["   Mantenga los encabezados de columna en la fila 1."],
    ]
    for row in instrucciones:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 55
    ws2.column_dimensions['B'].width = 45
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    wb.close()
    
    filename = f"CMMS-BioAI_Plantilla_Repuestos_{dt.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ---------------------------------------------------------
# ENDPOINT: IMPORTAR REPUESTOS DESDE EXCEL
# ---------------------------------------------------------
@router.post("/import-excel")
async def importar_repuestos_excel(file: UploadFile = File(...), session: Session = Depends(get_session)):
    """
    Importa repuestos desde un archivo Excel (.xlsx).
    Columnas esperadas (encabezado en fila 1):
      nombre_repuesto *, numero_material, descripcion,
      cantidad_disponible *, unidad_medida, ubicacion_almacen, nivel_stock_minimo
    
    - Campo obligatorio: nombre_repuesto, cantidad_disponible
    - Si numero_material ya existe, se ACTUALIZA el registro (upsert)
    """
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx o .csv")
    
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo no debe superar 5MB")
    
    # Determinar si es Excel o CSV
    ext = Path(file.filename).suffix.lower() if file.filename else ''
    is_csv = ext == '.csv'
    
    if is_csv:
        try:
            text = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                text = contents.decode('latin-1')
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error al decodificar el archivo CSV: {str(e)}")
        
        try:
            sniffer_sample = text[:2048]
            try:
                dialect = csv.Sniffer().sniff(sniffer_sample, delimiters=',;\t')
            except csv.Error:
                dialect = csv.excel
            
            reader = csv.reader(StringIO(text), dialect)
            filas = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo CSV: {str(e)}")
        
        if len(filas) < 2:
            raise HTTPException(status_code=400, detail="El archivo CSV est vacio o solo tiene encabezados")
    else:
        try:
            wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error al leer el archivo Excel: {str(e)}")
    
    COLUMNAS = [
        'nombre_repuesto', 'numero_serie', 'numero_material',
        'descripcion', 'especificaciones_tecnicas',
        'cantidad_disponible', 'unidad_medida', 'ubicacion_almacen', 'nivel_stock_minimo',
        'proveedor_ultimo', 'fecha_ultima_entrada', 'precio_referencia'
    ]
    OBLIGATORIAS = {'nombre_repuesto', 'cantidad_disponible'}
    
    # Buscar la hoja con encabezados (solo Excel)
    if not is_csv:
        ws = None
        for sheet in wb.worksheets:
            filas_tmp = list(sheet.iter_rows(values_only=True))
            if filas_tmp:
                enc_tmp = [str(c).strip().lower() if c else '' for c in filas_tmp[0]]
                if OBLIGATORIAS.issubset(set(enc_tmp)):
                    ws = sheet
                    break
        
        if ws is None:
            ws = wb.worksheets[0]
        
        filas = list(ws.iter_rows(values_only=True))
        if len(filas) < 2:
            raise HTTPException(status_code=400, detail="El archivo esta vacio o solo tiene encabezados")
    
    encabezados_leidos = [str(c).strip().lower() if c else '' for c in filas[0]]
    
    col_index = {}
    for i, h in enumerate(encabezados_leidos):
        if h in COLUMNAS:
            col_index[h] = i
    
    faltantes = OBLIGATORIAS - set(col_index.keys())
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas obligatorias: {', '.join(faltantes)}. Encabezados encontrados: {', '.join(encabezados_leidos)}"
        )
    
    exitosos = 0
    actualizados = 0
    fallidos = []
    
    for fila_num, fila in enumerate(filas[1:], start=2):
        try:
            def get_val(col_name):
                idx = col_index.get(col_name)
                if idx is None or idx >= len(fila):
                    return None
                val = fila[idx]
                return str(val).strip() if val is not None else None
            
            errores_fila = []
            for col in OBLIGATORIAS:
                val = get_val(col)
                if not val:
                    errores_fila.append(f"'{col}' es obligatorio")
            
            if errores_fila:
                fallidos.append({
                    "fila": fila_num,
                    "nombre": get_val('nombre_repuesto') or 'N/A',
                    "errores": errores_fila
                })
                continue
            
            # Parsear cantidad_disponible
            try:
                cantidad = int(float(get_val('cantidad_disponible')))
                if cantidad < 0:
                    raise ValueError()
            except (ValueError, TypeError):
                fallidos.append({
                    "fila": fila_num,
                    "nombre": get_val('nombre_repuesto') or 'N/A',
                    "errores": ["cantidad_disponible debe ser un número entero positivo"]
                })
                continue
            
            # Parsear nivel_stock_minimo (opcional)
            stock_min = None
            stock_str = get_val('nivel_stock_minimo')
            if stock_str:
                try:
                    stock_min = int(float(stock_str))
                except (ValueError, TypeError):
                    pass  # No es fatal, se ignora
            
            # Parsear precio_referencia (opcional)
            precio_ref = None
            precio_str = get_val('precio_referencia')
            if precio_str:
                try:
                    precio_ref = float(precio_str)
                except (ValueError, TypeError):
                    pass
            
            # Parsear fecha_ultima_entrada (opcional, formato YYYY-MM-DD)
            fecha_ent = None
            fecha_str = get_val('fecha_ultima_entrada')
            if fecha_str:
                try:
                    from datetime import date as date_mod
                    # Intentar varios formatos
                    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                        try:
                            fecha_ent = date_mod.strptime(str(fecha_str).strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass
            
            unidad = get_val('unidad_medida') or 'unidad'
            
            # Verificar si ya existe (por numero_material o nombre)
            num_mat = get_val('numero_material')
            rep_existente = None
            
            if num_mat:
                rep_existente = session.exec(
                    select(Repuesto).where(Repuesto.numero_material == num_mat)
                ).first()
            
            if not rep_existente:
                # También buscar por nombre como fallback
                nombre = get_val('nombre_repuesto')
                rep_existente = session.exec(
                    select(Repuesto).where(Repuesto.nombre_repuesto == nombre)
                ).first()
            
            if rep_existente:
                # ACTUALIZAR
                rep_existente.nombre_repuesto = get_val('nombre_repuesto') or rep_existente.nombre_repuesto
                rep_existente.numero_serie = get_val('numero_serie') or rep_existente.numero_serie
                rep_existente.numero_material = num_mat or rep_existente.numero_material
                rep_existente.descripcion = get_val('descripcion') or rep_existente.descripcion
                rep_existente.especificaciones_tecnicas = get_val('especificaciones_tecnicas') or rep_existente.especificaciones_tecnicas
                rep_existente.cantidad_disponible = cantidad
                rep_existente.unidad_medida = unidad
                rep_existente.ubicacion_almacen = get_val('ubicacion_almacen') or rep_existente.ubicacion_almacen
                rep_existente.nivel_stock_minimo = stock_min if stock_min is not None else rep_existente.nivel_stock_minimo
                rep_existente.proveedor_ultimo = get_val('proveedor_ultimo') or rep_existente.proveedor_ultimo
                rep_existente.fecha_ultima_entrada = fecha_ent or rep_existente.fecha_ultima_entrada
                rep_existente.precio_referencia = precio_ref if precio_ref is not None else rep_existente.precio_referencia
                session.add(rep_existente)
                actualizados += 1
            else:
                # CREAR nuevo
                nuevo = Repuesto(
                    nombre_repuesto=get_val('nombre_repuesto'),
                    numero_serie=get_val('numero_serie'),
                    numero_material=num_mat,
                    descripcion=get_val('descripcion'),
                    especificaciones_tecnicas=get_val('especificaciones_tecnicas'),
                    cantidad_disponible=cantidad,
                    unidad_medida=unidad,
                    ubicacion_almacen=get_val('ubicacion_almacen'),
                    nivel_stock_minimo=stock_min,
                    proveedor_ultimo=get_val('proveedor_ultimo'),
                    fecha_ultima_entrada=fecha_ent,
                    precio_referencia=precio_ref
                )
                session.add(nuevo)
                exitosos += 1
                
        except Exception as e:
            fallidos.append({
                "fila": fila_num,
                "nombre": get_val('nombre_repuesto') if col_index.get('nombre_repuesto') is not None else 'N/A',
                "errores": [str(e)]
            })
    
    session.commit()
    if not is_csv:
        wb.close()
    
    return {
        "exitosos": exitosos,
        "actualizados": actualizados,
        "fallidos": len(fallidos),
        "total_procesados": exitosos + actualizados + len(fallidos),
        "errores": fallidos
    }


# ============================================================
# RUTAS DINÁMICAS (van DESPUÉS de las estáticas)
# ============================================================

# ---------------------------------------------------------
# ENDPOINT: SUBIR IMAGEN PRINCIPAL DE REPUESTO
# Estructura: uploads/INVENTARIO/I0001_nombre/I0001_nombre.ext
# ---------------------------------------------------------
@router.post("/{rep_id}/upload_imagen")
async def upload_imagen_repuesto(rep_id: int, file: UploadFile = File(...), session: Session = Depends(get_session)):
    """Sube una imagen principal para un repuesto."""
    db_rep = session.get(Repuesto, rep_id)
    if not db_rep:
        raise HTTPException(status_code=404, detail="Repuesto no encontrado")

    allowed_types = ["image/jpeg", "image/png", "image/gif", "image/webp", "image/bmp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Use JPEG, PNG, GIF, WebP o BMP.")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="La imagen no debe superar 5MB")

    rep_code = f"I{rep_id:04d}"
    nombre_safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in (db_rep.nombre_repuesto or "SN"))
    folder_name = f"{rep_code}_{nombre_safe}"

    uploads_dir = get_dir("inventario_imagenes") / folder_name
    uploads_dir.mkdir(parents=True, exist_ok=True)

    ext = os.path.splitext(file.filename)[1] if file.filename else ".jpg"
    filename = f"{folder_name}{ext}"
    file_path = uploads_dir / filename

    # Eliminar imagen previa con otra extension
    for old_ext in [".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"]:
        old_path = uploads_dir / f"{folder_name}{old_ext}"
        if old_path.exists() and old_path != file_path:
            old_path.unlink()

    with open(str(file_path), "wb") as f:
        f.write(contents)

    imagen_ruta = f"INVENTARIO/{folder_name}/{filename}"
    db_rep.imagen_ruta = imagen_ruta
    session.add(db_rep)
    session.commit()
    session.refresh(db_rep)

    return {"ok": True, "imagen_ruta": db_rep.imagen_ruta, "nombre_archivo": file.filename}


# ---------------------------------------------------------
# ENDPOINT: ELIMINAR IMAGEN PRINCIPAL DE REPUESTO
# ---------------------------------------------------------
@router.delete("/{rep_id}/imagen")
def eliminar_imagen_repuesto(rep_id: int, session: Session = Depends(get_session)):
    """Elimina la imagen principal de un repuesto (archivo + BD)."""
    db_rep = session.get(Repuesto, rep_id)
    if not db_rep:
        raise HTTPException(status_code=404, detail="Repuesto no encontrado")

    if not db_rep.imagen_ruta:
        raise HTTPException(status_code=400, detail="Este repuesto no tiene imagen")

    file_path = get_dir("uploads_base") / db_rep.imagen_ruta
    if file_path.exists():
        file_path.unlink()

    db_rep.imagen_ruta = None
    session.add(db_rep)
    session.commit()

    return {"ok": True, "message": "Imagen eliminada"}


@router.get("/{rep_id}", response_model=RepuestoRead)
def obtener_repuesto(rep_id: int, session: Session = Depends(get_session)):
    db_rep = session.get(Repuesto, rep_id)
    if not db_rep:
        raise HTTPException(status_code=404, detail="Repuesto no encontrado")
    return db_rep


@router.put("/{rep_id}", response_model=RepuestoRead)
def actualizar_repuesto(
    rep_id: int,
    datos: RepuestoUpdate,
    session: Session = Depends(get_session),
):
    db_rep = session.get(Repuesto, rep_id)
    if not db_rep:
        raise HTTPException(status_code=404, detail="Repuesto no encontrado")
    for key, value in datos.model_dump(exclude_unset=True).items():
        setattr(db_rep, key, value)
    session.add(db_rep)
    session.commit()
    session.refresh(db_rep)
    return db_rep


@router.delete("/{rep_id}")
def eliminar_repuesto(rep_id: int, session: Session = Depends(get_session)):
    db_rep = session.get(Repuesto, rep_id)
    if not db_rep:
        raise HTTPException(status_code=404, detail="Repuesto no encontrado")
    session.delete(db_rep)
    session.commit()
    return {"ok": True, "message": "Repuesto eliminado"}
