"""
Genera los 8 archivos estáticos de plantillas para CMMS-BioAI:
- 4 archivos Excel (.xlsx): Equipos, Repuestos, Herramientas, Proveedores
- 4 archivos CSV (.csv): Equipos, Repuestos, Herramientas, Proveedores

Los archivos se colocan en:
    /home/z/my-project/cmmsbioai/frontend/public/plantillas/

Una vez allí, Vite los sirve directamente en la URL /plantillas/<nombre>
y el frontend puede descargarlos sin llamar al backend.

Los datos demo son los mismos que tenían los endpoints dinámicos del backend,
así que el comportamiento para el usuario es idéntico.
"""
import openpyxl
import csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

OUTPUT_DIR = Path("/home/z/my-project/cmmsbioai/frontend/public/plantillas")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Estilos comunes para Excel
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def crear_excel(nombre_archivo: str, hoja_titulo: str, color_header: str,
                encabezados: list, datos_demo: list, anchos: list,
                instrucciones: list):
    """Crea un archivo Excel con encabezados, datos demo y hoja de instrucciones."""
    path = OUTPUT_DIR / nombre_archivo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja_titulo

    # Encabezados
    ws.append(encabezados)
    header_fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    for col_num, _ in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Datos demo
    for fila in datos_demo:
        ws.append(fila)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(encabezados)):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN

    # Anchos
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    ws.row_dimensions[1].height = 32

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    for row in instrucciones:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 65
    ws2.column_dimensions['B'].width = 55

    # Metadata
    wb.properties.creator = "CMMS-BioAI"
    wb.properties.title = f"Plantilla {hoja_titulo}"

    wb.save(path)
    wb.close()
    print(f"  ✅ {path.name} ({path.stat().st_size} bytes)")


def crear_csv(nombre_archivo: str, encabezados: list, datos_demo: list):
    """Crea un archivo CSV con encabezados y datos demo."""
    path = OUTPUT_DIR / nombre_archivo
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(encabezados)
        for fila in datos_demo:
            writer.writerow(fila)
    print(f"  ✅ {path.name} ({path.stat().st_size} bytes)")


# ============================================================
# 1. EQUIPOS
# ============================================================
print("\n=== EQUIPOS ===")
EQ_HEADERS = [
    'nombre_corto', 'modelo', 'numero_serie', 'numero_material', 'marca', 'fecha_adquisicion',
    'fecha_fin_garantia', 'ubicacion_actual', 'estado', 'proveedor_principal', 'registro_sanitario_bolivia',
    'descripcion', 'calibracion_proxima', 'responsable_username'
]
EQ_DATOS = [
    ["Microscopio Olympus CX23", "CX23", "MIC-OLY-001", "MAT-CX23-A", "Olympus", "2023-03-15", "2025-03-15", "Lab. Microbiología", "Operativo", "Olympus Bolivia", "RS-BOL-2023-001", "Microscopio binocular para microbiología clínica", "2025-03-15", ""],
    ["Centrífuga Eppendorf 5424", "5424", "CEN-EPP-002", "MAT-5424-B", "Eppendorf", "2022-07-20", "2024-07-20", "Lab. Hematología", "Operativo", "Eppendorf Latam", "RS-BOL-2022-045", "Centrífuga de mesa 24 tubos, rotor FA-45-24-11", "2024-07-20", ""],
    ["Autoclave Steris Amsco 3043", "3043", "AUT-STR-003", "", "Steris", "2021-01-10", "2023-01-10", "Central de Esterilización", "Operativo", "Steris Corporation", "", "Autoclave hospitalaria vertical, 400L capacidad", "2025-01-10", ""],
    ["Analizador Bioquímico Cobas c311", "c311", "ANA-ROC-004", "MAT-C311-V2", "Roche", "2023-06-01", "2025-06-01", "Lab. Bioquímica", "Operativo", "Roche Diagnostics Bolivia", "RS-BOL-2023-089", "Analizador bioquímico automatizado, 120 pruebas/h", "2025-06-01", ""],
    ["Monitor de Signos Vitales Mindray", "uMEC10", "MON-MIN-005", "", "Mindray", "2022-11-05", "2024-11-05", "UCI Box 3", "En Mantenimiento", "Mindray Bolivia", "RS-BOL-2022-112", "Monitor multiparámetro con SpO2, ECG, NIBP, Temp", "2024-11-05", ""],
    ["Electrocardiógrafo GE MAC 2000", "MAC 2000", "ELE-GEE-006", "", "GE Healthcare", "2020-09-15", "2022-09-15", "Cardiología", "Operativo", "GE Healthcare Latam", "RS-BOL-2020-034", "ECG 12 derivaciones con interpretación", "2024-09-15", ""],
    ["Espectrofotómetro Thermo Genesys 30", "Genesys 30", "ESP-THM-007", "MAT-GEN30", "Thermo Fisher", "2023-02-28", "2025-02-28", "Lab. Investigación", "Operativo", "Thermo Fisher Scientific", "", "Espectrofotómetro visible, rango 325-1100nm", "2025-02-28", ""],
    ["Balanza Analítica Sartorius", "Quintix 224", "BAL-SAR-008", "", "Sartorius", "2021-08-12", "2023-08-12", "Lab. Control Calidad", "Operativo", "Sartorius Bolivia", "", "Balanza analítica 220g / 0.1mg, calibración interna", "2024-08-12", ""],
    ["Desfibrilador Zoll R Series", "R Series", "DES-ZOL-009", "MAT-RS-PRO", "Zoll", "2022-04-18", "2024-04-18", "Emergencias", "Operativo", "Zoll Medical", "RS-BOL-2022-078", "Desfibrilador biphasic con monitor y marcapasos", "2025-04-18", ""],
    ["Incubadora CO2 Thermo Heracell", "Heracell VIOS 160i", "INC-THM-010", "", "Thermo Fisher", "2023-09-01", "2025-09-01", "Lab. Cultivo Celular", "Fuera de Servicio", "Thermo Fisher Scientific", "", "Incubadora CO2 con control de O2, 170L", "2025-09-01", ""],
    ["Bomba de Infusión B. Braun", "Infusomat Space", "BOM-BRA-011", "MAT-INF-SP2", "B. Braun", "2022-12-10", "2024-12-10", "UCI Box 1", "Operativo", "B. Braun Bolivia", "RS-BOL-2022-156", "Bomba de infusión volumétrica con modo PCA", "2024-12-10", ""],
    ["Termociclador Bio-Rad T100", "T100", "TER-BIO-012", "", "Bio-Rad", "2021-05-22", "2023-05-22", "Lab. Biología Molecular", "Operativo", "Bio-Rad Latam", "", "Termociclador PCR 96 pocillos, gradiente", "2025-05-22", ""],
    ["Respirador Dräger Evita V500", "Evita V500", "RES-DRA-013", "", "Dräger", "2023-01-15", "2025-01-15", "UCI Box 5", "En Mantenimiento", "Dräger Bolivia", "RS-BOL-2023-003", "Ventilador de UCI con modos avanzados", "2025-01-15", ""],
    ["Pipeta Automática Hamilton", "Microlab STARlet", "PIP-HAM-014", "MAT-STAR-V3", "Hamilton", "2022-03-08", "2024-03-08", "Lab. Automatización", "Operativo", "Hamilton Company", "", "Pipeta automatizada 8 canales, 1-1000uL", "2025-03-08", ""],
    ["Cámara de Flujo Laminar Esco", "A2 Class", "CAM-ESC-015", "", "Esco", "2020-06-25", "2022-06-25", "Lab. Microbiología", "Operativo", "Esco Technologies", "", "Cabina de seguridad biológica Clase II Tipo A2", "2024-06-25", ""],
]
EQ_ANCHOS = [28, 18, 18, 16, 18, 18, 18, 24, 18, 24, 24, 40, 18, 22]
EQ_INSTR = [
    ["INSTRUCCIONES PARA IMPORTAR EQUIPOS"],
    [""],
    ["1. Los campos marcados como obligatorios (*) no pueden estar vacíos:"],
    ["   - modelo *", "modelo del equipo"],
    ["   - numero_serie *", "debe ser único, si ya existe se ACTUALIZARÁ el registro"],
    ["   - marca *", "fabricante del equipo"],
    ["   - fecha_adquisicion *", "formato: YYYY-MM-DD o DD/MM/YYYY"],
    [""],
    ["2. Campo 'estado': debe coincidir con un estado existente en el sistema."],
    ["   Estados típicos: Operativo, En Mantenimiento, Fuera de Servicio"],
    ["   Si no coincide, se usará el estado por defecto."],
    [""],
    ["3. Campo 'responsable_username': username del usuario registrado en el sistema."],
    ["   Si no coincide, se asignará sin responsable."],
    [""],
    ["4. Campo 'calibracion_proxima': fecha en formato YYYY-MM-DD (opcional)."],
    [""],
    ["5. Límite de archivo: 5MB (~1000 equipos). Solo archivos .xlsx o .csv."],
    [""],
    ["6. Si un 'numero_serie' ya existe, el equipo se ACTUALIZARÁ con los nuevos datos."],
    [""],
    ["7. Puede eliminar los datos de ejemplo y usar sus propios datos."],
    ["   Mantenga los encabezados de columna en la fila 1."],
]
crear_excel("plantilla_equipos.xlsx", "Equipos CMMS-BioAI", "2C3E50", EQ_HEADERS, EQ_DATOS, EQ_ANCHOS, EQ_INSTR)
crear_csv("plantilla_equipos.csv", EQ_HEADERS, EQ_DATOS)


# ============================================================
# 2. REPUESTOS
# ============================================================
print("\n=== REPUESTOS ===")
REP_HEADERS = [
    'nombre_repuesto', 'numero_serie', 'numero_material',
    'descripcion', 'especificaciones_tecnicas',
    'cantidad_disponible', 'unidad_medida', 'ubicacion_almacen', 'nivel_stock_minimo',
    'proveedor_ultimo', 'fecha_ultima_entrada', 'precio_referencia'
]
REP_DATOS = [
    ["Filtro HEPA para cabina de flujo", "SN-FLT-001", "FLT-HEP-001", "Filtro HEPA 99.97% 0.3μm para cabina Esco A2", "99.97% 0.3um, tamano estandar", 3, "unidad", "Almacén B - Estante 1", 2, "BioSupply SRL", "2025-01-15", 850.50],
    ["Lámpara UV para autoclave", "SN-LMP-002", "LMP-UV-002", "Lámpara germicida UV 254nm Steris", "254nm, 15W, T8", 5, "unidad", "Almacén B - Estante 1", 2, "MedEquip Bolivia", "2025-02-20", 320.00],
    ["Electrodo pH recalculado", "SN-ELC-003", "ELC-PH-003", "Electrodo de pH combinado para analizadores", "Combinado, rango 0-14 pH", 4, "unidad", "Lab. Bioquímica", 2, "LabTech SA", "2025-03-10", 1200.00],
    ["Tubos de ensayo 16x100mm", "", "TUB-16004", "Tubos de vidrio borosilicato, paquete x100", "16x100mm, borosilicato", 25, "paquete", "Almacén A - Estante 3", 10, "Vidriolab", "2025-04-01", 45.00],
    ["Reactivos panel bioquímico", "", "RCT-BIO-005", "Kit de reactivos Cobas c311 panel completo", "Panel completo, vigencia 6 meses", 2, "kit", "Refrigerador Lab 2°C-8°C", 1, "Roche Diagnostics", "2025-05-15", 3500.00],
    ["Filtro de aire compresor", "", "FLT-AIR-006", "Filtro de línea para compresor dental/medico", "5 micras, roscable 1/4", 6, "unidad", "Almacén B - Estante 2", 3, "DentalPro", "2025-03-22", 120.00],
    ["Cable ECG 12 derivaciones", "SN-CBL-007", "CBL-ECG-007", "Cable paciente 12-lead para GE MAC 2000", "12-lead, conector GE", 3, "unidad", "Almacén C - Gabinete 1", 2, "GE Healthcare", "2025-01-30", 2100.00],
    ["Sensor SpO2 dedo", "", "SNR-SPO-008", "Sensor pulsioximetría dedo adulto Mindray", "Dedo adulto, compatible Mindray", 4, "unidad", "UCI - Armario suministros", 2, "Mindray", "2025-04-10", 780.00],
    ["Papel térmico monitor", "", "PPL-TRM-009", "Papel térmico 112mm para monitor de signos vitales", "112mm x 30m", 8, "rollo", "Almacén A - Estante 5", 3, "PaperMed", "2025-06-01", 25.00],
    ["Batería interna desfibrilador", "SN-BAT-010", "BAT-DEF-010", "Batería recargable Zoll R Series 14.4V", "14.4V Li-Ion, autonomia 8h", 2, "unidad", "Emergencias - Armario", 1, "Zoll Medical", "2024-12-15", 4500.00],
    ["Goma/pistón pipeta", "", "GOM-PIP-011", "Pistón de goma para pipeta Hamilton 100-1000μL", "Compatible Hamilton 100-1000uL", 10, "unidad", "Lab. Automatización", 5, "Hamilton", "2025-05-20", 35.00],
    ["Sellador de tapa incubadora", "", "SLR-INC-012", "Junta de goma puerta incubadora Thermo Heracell", "Silicona, tamano estandar Heracell", 2, "unidad", "Almacén B - Estante 4", 1, "Thermo Fisher", "2025-02-28", 290.00],
    ["Fluorescente microscopio", "", "LMP-MIC-013", "Lámpara halógena 6V 30W para Olympus CX23", "6V 30W, base BA15d", 3, "unidad", "Almacén B - Estante 1", 2, "Olympus", "2025-03-15", 180.00],
    ["Reactivos hematología", "", "RCT-HEM-014", "Kit diluyente/limpiador analizador hematológico", "Diluyente + limpiador + hemolizante", 3, "kit", "Refrigerador Lab 2°C-8°C", 1, "Sysmex", "2025-04-20", 2800.00],
    ["Filtro CO2 incubadora", "", "FLT-CO2-015", "Filtro HEPA para incubadora CO2 Thermo", "HEPA, 99.97% para Thermo 150/250", 2, "unidad", "Almacén B - Estante 3", 1, "Thermo Fisher", "2025-01-25", 650.00],
    ["Aceite bomba vacío", "", "ACE-VAC-016", "Aceite para bomba de vacío Autoclave Steris 1L", "1L, grado vacío", 4, "litro", "Almacén A - Estante 2", 2, "Steris", "2025-05-05", 150.00],
    ["Manguera silicona", "", "MNG-SIL-017", "Tubo silicona 8mm ID para bombas de infusión", "8mm ID, medical grade", 5, "metro", "Almacén C - Estante 1", 3, "MedSilicone", "2025-04-15", 18.00],
    ["Electrodo ECG desechable", "", "ELC-ECG-018", "Electrodos Ag/Cl adulto paquete x50", "Ag/Cl, adulto, x50", 20, "paquete", "UCI - Armario suministros", 10, "3M Healthcare", "2025-06-10", 55.00],
    ["Papel filtro centrífuga", "", "PPL-CEN-019", "Filtro de carbón para Eppendorf 5424", "Carbón activado, compatible 5424", 6, "unidad", "Lab. Hematología", 3, "Eppendorf", "2025-03-30", 95.00],
    ["Solución calibración pH", "", "SLC-PH-020", "Solución buffer pH 4.01 / 7.00 / 10.01 set", "Set 3 botellas 500mL", 5, "kit", "Lab. Control Calidad", 2, "Hanna Instruments", "2025-02-10", 280.00],
]
REP_ANCHOS = [35, 16, 18, 50, 30, 14, 12, 28, 18, 22, 18, 14]
REP_INSTR = [
    ["INSTRUCCIONES PARA IMPORTAR REPUESTOS"],
    [""],
    ["1. Los campos marcados como obligatorios (*) no pueden estar vacíos:"],
    ["   - nombre_repuesto *", "nombre del repuesto o insumo"],
    ["   - cantidad_disponible *", "número entero (ej: 5)"],
    [""],
    ["2. Campo 'numero_material': código del repuesto. Si ya existe, se ACTUALIZARÁ."],
    ["   Si no existe numero_material, se busca por nombre_repuesto."],
    [""],
    ["3. Campo 'numero_serie': número de serie del repuesto (opcional)."],
    ["4. Campo 'especificaciones_tecnicas': voltaje, tamaño, capacidad, etc. (opcional)."],
    ["5. Campo 'nivel_stock_minimo': entero. Cantidad mínima antes de alerta."],
    ["   Si no se especifica, no se generará alerta de stock bajo."],
    [""],
    ["6. Campo 'unidad_medida': unidad, par, metro, litro, kit, paquete, rollo, etc."],
    ["   Si se deja vacío, se usará 'unidad' por defecto."],
    [""],
    ["7. Campo 'proveedor_ultimo': nombre del último proveedor (opcional)."],
    ["8. Campo 'fecha_ultima_entrada': fecha en formato YYYY-MM-DD (opcional)."],
    ["9. Campo 'precio_referencia': precio en Bs. número decimal (opcional)."],
    [""],
    ["10. Límite de archivo: 5MB. Solo archivos .xlsx o .csv."],
    [""],
    ["11. Puede eliminar los datos de ejemplo y usar sus propios datos."],
    ["   Mantenga los encabezados de columna en la fila 1."],
]
crear_excel("plantilla_repuestos.xlsx", "Repuestos CMMS-BioAI", "2C3E50", REP_HEADERS, REP_DATOS, REP_ANCHOS, REP_INSTR)
crear_csv("plantilla_repuestos.csv", REP_HEADERS, REP_DATOS)


# ============================================================
# 3. HERRAMIENTAS
# ============================================================
print("\n=== HERRAMIENTAS ===")
HER_HEADERS = [
    'nombre_herramienta', 'numero_identificacion', 'descripcion', 'categoria',
    'cantidad_disponible', 'unidad_medida', 'ubicacion_almacen', 'estado_uso',
    'costo_adquisicion', 'fecha_adquisicion', 'proveedor_ultimo', 'observaciones'
]
HER_DATOS = [
    ["Osciloscopio digital Rigol DS1054Z", "OSC-001", "Osciloscopio 4 canales 50MHz para diagnostico de equipos medicos", "Instrumento de Medición", 1, "unidad", "Taller - Estante A1", "Disponible", 5200.00, "2024-03-15", "Rigol Technologies", "Calibracion anual vigente"],
    ["Multímetro Fluke 87V", "MUL-002", "Multímetro digital True RMS industrial", "Instrumento de Medición", 2, "unidad", "Taller - Estante A1", "Disponible", 2800.00, "2024-01-20", "Fluke Corporation", "Unidad con bateria nueva"],
    ["Analizador de seguridad eléctrica Fluke ESA620", "ANA-003", "Para prueba de seguridad eléctrica en equipos médicos según IEC 62353", "Instrumento de Medición", 1, "unidad", "Taller - Estante A2", "En Uso", 15000.00, "2024-06-10", "Fluke Corporation", "En uso en mantenimiento preventivo"],
    ["Juego de destornilladores aislados Wiha", "DES-004", "Juego 12 piezas VDE 1000V para trabajo en equipos energizados", "Herramienta Manual", 3, "juego", "Taller - Cajón B1", "Disponible", 450.00, "2024-02-05", "Wiha", ""],
    ["Pinza amperimérica Fluke 376 FC", "PIN-005", "Pinza AC/DC 1000A con Bluetooth", "Instrumento de Medición", 1, "unidad", "Taller - Estante A1", "Disponible", 3500.00, "2024-04-18", "Fluke Corporation", "Con certificado de calibracion"],
    ["Alcohol isopropílico 99%", "ALC-006", "Para limpieza de componentes electrónicos y ópticos", "Consumible", 5, "litro", "Taller - Estante C2", "Disponible", 85.00, "2025-01-10", "Quimica Bolivia", "Almacenar en lugar ventilado"],
    ["Estaño para soldadura 60/40 0.8mm", "EST-007", "Rollo 500g soldadura electrónica con núcleo de resina", "Consumible", 4, "rollo", "Taller - Estante C1", "Disponible", 120.00, "2025-02-15", "ElectroSolder", ""],
    ["Kit de calibración presión", "KIT-008", "Kit con manómetro patrón y accesorios para calibración de equipos de presión", "Kit", 1, "kit", "Taller - Estante A3", "Disponible", 8500.00, "2024-08-01", "Additel", "Incluye case y certificados"],
    ["Generador de señales Rigol DG822", "GEN-009", "Generador de funciones 2 canales 30MHz para pruebas de equipos", "Instrumento de Medición", 1, "unidad", "Taller - Estante A2", "Disponible", 4100.00, "2024-05-22", "Rigol Technologies", ""],
    ["Cautín soldador Weller WE1010", "CAU-010", "Estación de soldadura 70W con control de temperatura", "Herramienta Manual", 2, "unidad", "Taller - Estante B1", "Disponible", 680.00, "2024-09-12", "Weller", "Una unidad en reparacion"],
]
HER_ANCHOS = [40, 18, 55, 24, 14, 12, 24, 16, 14, 16, 22, 35]
HER_INSTR = [
    ["INSTRUCCIONES PARA IMPORTAR HERRAMIENTAS"],
    [""],
    ["1. Los campos marcados como obligatorios (*) no pueden estar vacíos:"],
    ["   - nombre_herramienta *", "nombre de la herramienta o material"],
    ["   - cantidad_disponible *", "número entero (ej: 1)"],
    [""],
    ["2. Campo 'categoria' (valores permitidos):"],
    ["   - Instrumento de Medición", "osciloscopios, multímetros, analizadores, etc."],
    ["   - Herramienta Manual", "destornilladores, alicates, soldadores, etc."],
    ["   - Consumible", "alcohol, estaño, cintas, limpiadores, etc."],
    ["   - Kit", "conjuntos de herramientas o accesorios"],
    [""],
    ["3. Campo 'estado_uso' (valores permitidos):"],
    ["   - Disponible", "lista para usar"],
    ["   - En Uso", "actualmente en uso en una OT o actividad"],
    ["   - En Reparación", "fuera de servicio temporalmente"],
    ["   - Dado de Baja", "retirada del servicio permanentemente"],
    [""],
    ["4. Campo 'numero_identificacion': código interno del taller (opcional pero recomendado)."],
    ["5. Campo 'costo_adquisicion': precio en Bs. número decimal (opcional)."],
    ["6. Campo 'fecha_adquisicion': fecha en formato YYYY-MM-DD (opcional)."],
    [""],
    ["7. Límite de archivo: 5MB. Solo archivos .xlsx o .csv."],
    [""],
    ["8. Puede eliminar los datos de ejemplo y usar sus propios datos."],
    ["   Mantenga los encabezados de columna en la fila 1."],
]
crear_excel("plantilla_herramientas.xlsx", "Herramientas CMMS-BioAI", "1B4332", HER_HEADERS, HER_DATOS, HER_ANCHOS, HER_INSTR)
crear_csv("plantilla_herramientas.csv", HER_HEADERS, HER_DATOS)


# ============================================================
# 4. PROVEEDORES
# ============================================================
print("\n=== PROVEEDORES ===")
PROV_HEADERS = [
    'nombre_empresa', 'ciudad', 'direccion',
    'telefono_principal', 'email_principal', 'pagina_web', 'notas_generales'
]
PROV_DATOS = [
    ["TechMed Bolivia SRL", "Cochabamba", "Av. Blanco Galindo km 7.5", "+591 4 4223344", "ventas@techmed.bo", "https://techmed.bo", "Distribuidor autorizado de equipos de monitoreo y signos vitales. Garantía 1 año en todos los equipos."],
    ["BioSupply Andina", "La Paz", "Calle Mercado N° 1234, Zona Sopocachi", "+591 2 2445566", "info@biosupply.bo", "https://biosupply.bo", "Repuestos y consumibles para laboratorio clínico. Descuento volumen 10% en pedidos >5000 Bs."],
    ["MedEquip Bolivia SA", "Santa Cruz", "Tercer Anillo Interno, Av. Roca y Coronado", "+591 3 3467788", "ventas@medequip.bo", "", "Importador de equipos de imagenología (ecógrafos, rayos X). Servicio técnico propio."],
    ["LabTech SRL", "Cochabamba", "Av. América Esq. Esteban Arze", "+591 4 4521199", "contacto@labtech.bo", "https://labtech.bo", "Especialistas en equipos de laboratorio: centrífugas, espectrofotómetros, balanzas analíticas."],
    ["Hospimed Bolivia", "La Paz", "Av. 6 de Agosto N° 2115", "+591 2 2150500", "ventas@hospimed.bo", "https://hospimed.bo", "Equipos de terapia intensiva y monitoreo multiparámetro. Servicio 24/7 en UCI."],
    ["DentalPro Suministros", "Cochabamba", "Calle Juan Caprario N° 0456", "+591 4 4256677", "info@dentalpro.bo", "https://dentalpro.bo", "Unidades dentales, autoclaves y consumibles. Cursos de capacitación incluidos."],
    ["Roche Diagnostics Bolivia", "La Paz", "Av. Mariscal Santa Cruz, Edif. Cámara Nacional", "+591 2 2334455", "bolivia@roche.com", "https://diagnostics.roche.com", "Reactivos y equipos de diagnóstico in vitro. Soporte técnico internacional."],
    ["Thermo Fisher Andina", "Santa Cruz", "Parque Industrial Mz 7 Lote 12", "+591 3 3528899", "andina@thermofisher.com", "https://thermofisher.com", "Incubadoras, refrigeradores de laboratorio, centrífugas. Envíos directos desde EEUU."],
    ["Steris Bolivia SA", "Cochabamba", "Av. Petrolera km 4.5", "+591 4 4710022", "servicio@steris.bo", "https://steris.bo", "Esterilizadores y autoclaves hospitalarias. Mantenimiento preventivo trimestral."],
    ["Olympus Latin America", "La Paz", "Av. Arce N° 2798, Edif. Multicentro", "+591 2 2773344", "latam@olympus.com", "https://olympus-latinamerica.com", "Microscopios, endoscopios y equipamiento óptico médico. Garantía internacional 2 años."],
    ["Mindray Bolivia", "Santa Cruz", "Av. San Martín N° 1700", "+591 3 3391212", "bolivia@mindray.com", "https://mindray.com", "Monitores de pacientes, ventiladores y equipos de ultrasonido. Capacitación gratuita."],
    ["GE Healthcare Bolivia", "Cochabamba", "Av. Oquendo N° 1245", "+591 4 4533344", "bolivia@gehealthcare.com", "https://gehealthcare.com", "Equipos de imágenes médicas: TAC, MRI, rayos X digitales. Contratos de servicio anuales."],
    ["Zoll Medical Andina", "La Paz", "Calle Colón N° 1590", "+591 2 2314567", "andina@zoll.com", "https://zoll.com", "Desfibriladores, monitores de reanimación y equipos de emergencia. Curso ACLS incluido."],
    ["Hamilton Bolivia", "Cochabamba", "Av. Blanco Galindo N° 8890", "+591 4 4719988", "bolivia@hamiltoncompany.com", "https://hamiltoncompany.com", "Pipetas, dispensadores y consumibles de precisión. Calibración GRATIS primer año."],
    ["Eppendorf Andina", "Santa Cruz", "Av. Monseñor Rivero N° 550", "+591 3 3332211", "andina@eppendorf.com", "https://eppendorf.com", "Microcentrífugas, pipetas, tubos y consumibles de laboratorio. Garantía 3 años."],
    ["Hanna Instruments Bolivia", "La Paz", "Av. 16 de Julio N° 1730", "+591 2 2288776", "bolivia@hannainst.com", "https://hannainst.com", "Medidores de pH, conductividad y oxígeno disuelto. Soluciones buffer certificadas."],
    ["3M Healthcare Bolivia", "Cochabamba", "Av. Cristo Redentor N° 0990", "+591 4 4785566", "bolivia@3m.com", "https://3m.com", "Electrodos ECG, productos de esterilización y materiales médicos. Distribuidor oficial."],
    ["Vidriolab SRL", "Cochabamba", "Calle Sucre N° 0050", "+591 4 4513344", "ventas@vidriolab.bo", "", "Vidriería de laboratorio: tubos, matraces, pipetas de vidrio. Fabricación a pedido."],
    ["Sysmex Latin America", "La Paz", "Av. Ballivián N° 700", "+591 2 2126677", "latam@sysmex.com", "https://sysmex.com", "Analizadores hematológicos y reactivos. Contrato de mantenimiento preventivo anual."],
    ["PaperMed Bolivia", "Santa Cruz", "Av. Argentida N° 2320", "+591 3 3367788", "info@papermed.bo", "", "Papel térmico para monitores, formularios continuos y consumibles. Entrega inmediata."],
]
PROV_ANCHOS = [28, 14, 45, 16, 28, 28, 60]
PROV_INSTR = [
    ["INSTRUCCIONES PARA IMPORTAR PROVEEDORES"],
    [""],
    ["1. Campo obligatorio (*) - no puede estar vacío:"],
    ["   - nombre_empresa *", "nombre legal o comercial del proveedor (debe ser único)"],
    [""],
    ["2. Campo 'ciudad': ciudad principal donde opera el proveedor (ej: La Paz, Cochabamba, Santa Cruz)."],
    ["   Es independiente de 'direccion' y se usa para filtros."],
    [""],
    ["3. Campo 'direccion': dirección física completa de la empresa (calle, número, zona, etc.)."],
    ["4. Campo 'telefono_principal': teléfono de la empresa (con código de país +591 si aplica)."],
    ["5. Campo 'email_principal': correo electrónico principal de la empresa."],
    ["6. Campo 'pagina_web': URL del sitio web oficial (opcional)."],
    ["7. Campo 'notas_generales': observaciones comerciales, garantías, condiciones, etc."],
    [""],
    ["8. Si el 'nombre_empresa' ya existe en la base de datos, se ACTUALIZARÁ el registro (upsert)."],
    ["   Si no existe, se creará un nuevo proveedor."],
    [""],
    ["9. Los CONTACTOS asociados no se importan desde este Excel. Una vez importados los"],
    ["   proveedores, agrega contactos individualmente desde la página de Proveedores."],
    [""],
    ["10. Límite de archivo: 5MB. Solo archivos .xlsx o .csv."],
    [""],
    ["11. Puede eliminar los datos de ejemplo y usar sus propios datos."],
    ["    Mantenga los encabezados de columna en la fila 1."],
]
crear_excel("plantilla_proveedores.xlsx", "Proveedores CMMS-BioAI", "2C3E50", PROV_HEADERS, PROV_DATOS, PROV_ANCHOS, PROV_INSTR)
crear_csv("plantilla_proveedores.csv", PROV_HEADERS, PROV_DATOS)


# ============================================================
# RESUMEN
# ============================================================
print("\n" + "=" * 60)
print("✅ Plantillas estáticas generadas en:")
print(f"   {OUTPUT_DIR}")
print("=" * 60)
print("\nArchivos creados:")
for f in sorted(OUTPUT_DIR.iterdir()):
    print(f"  {f.name:40s} {f.stat().st_size:>8,} bytes")
