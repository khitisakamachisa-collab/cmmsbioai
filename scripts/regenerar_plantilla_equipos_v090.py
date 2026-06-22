"""
Regenera la plantilla Excel estática de equipos para v0.9.0.

Cambios vs v0.8.3:
- Encabezados actualizados (sin campos obsoletos, con campos nuevos)
- 2 filas de ejemplo (plantilla vacía según filosofía v0.9.0)
- Hoja de instrucciones actualizada para v0.9.0

Archivo destino: frontend/public/plantillas/plantilla_equipos.xlsx
"""
import openpyxl
import csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

OUTPUT_DIR = Path("/home/z/my-project/cmmsbioai/frontend/public/plantillas")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Estilos
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============================================================
# 1. EQUIPOS — v0.9.0
# ============================================================
print("=== Generando plantilla_equipos.xlsx (v0.9.0) ===")

EQ_HEADERS = [
    'nombre_corto', 'modelo', 'numero_serie', 'numero_material', 'marca',
    'fecha_adquisicion', 'fecha_inicio_garantia', 'fecha_fin_garantia',
    'ubicacion_actual', 'estado', 'proveedor_principal', 'condicion_origen',
    'descripcion', 'observaciones'
]

EQ_DATOS = [
    ["Microscopio Olympus CX23", "CX23", "MIC-OLY-001", "MAT-CX23-A", "Olympus",
     "2023-03-15", "2023-03-15", "2025-03-15", "Lab. Microbiología", "Operativo",
     "Olympus Bolivia", "Compra",
     "Microscopio binocular para microbiología clínica", "Funciona correctamente"],

    ["Monitor Signos Vitales Mindray", "uMEC10", "MON-MIN-002", "", "Mindray",
     "2022-11-05", "2022-11-05", "2024-11-05", "UCI Box 3", "En Mantenimiento",
     "Mindray Bolivia", "Donación",
     "Monitor multiparámetro con SpO2, ECG, NIBP, Temp", "Requiere calibración de SpO2"],
]

EQ_ANCHOS = [28, 18, 18, 16, 18, 18, 18, 18, 24, 18, 24, 18, 40, 35]

EQ_INSTR = [
    ["INSTRUCCIONES PARA IMPORTAR EQUIPOS v0.9.0"],
    [""],
    ["1. Campos obligatorios (*) no pueden estar vacíos:"],
    ["   - nombre_corto *", "nombre descriptivo del equipo (alias)"],
    ["   - modelo *", "modelo del equipo (NO editable después de creado)"],
    ["   - numero_serie *", "único, si ya existe se ACTUALIZA el registro (upsert)"],
    ["   - marca *", "fabricante (NO editable después de creado)"],
    [""],
    ["2. Campo 'estado': debe coincidir con un estado existente en el sistema."],
    ["   Estados típicos: Operativo, En Mantenimiento, Fuera de Servicio, etc."],
    ["   Si no coincide, se usará el estado por defecto (Operativo)."],
    [""],
    ["3. Campo 'proveedor_principal': nombre del proveedor."],
    ["   - Si el proveedor NO existe en la BD, se CREA automáticamente"],
    ["   - Luego podrá completar sus datos (ciudad, contacto, etc.) en la página Proveedores"],
    [""],
    ["4. Campo 'condicion_origen' (valores permitidos):"],
    ["   Compra, Donación, Préstamo, Demostración, Evaluación, Leasing, Renta, Comodato, Otro"],
    [""],
    ["5. Fechas: formato YYYY-MM-DD o DD/MM/YYYY."],
    ["   - fecha_adquisicion: opcional"],
    ["   - fecha_inicio_garantia: opcional (debe ser <= fecha_fin_garantia)"],
    ["   - fecha_fin_garantia: opcional"],
    [""],
    ["6. Campos 'descripcion' y 'observaciones':"],
    ["   - descripcion: descripción TÉCNICA del equipo (¿qué es? ¿qué hace?)"],
    ["   - observaciones: notas OPERATIVAS (¿cómo está? accesorios, advertencias)"],
    [""],
    ["7. Límite de archivo: 5MB. Solo archivos .xlsx o .csv."],
    [""],
    ["8. CAMPOS ELIMINADOS en v0.9.0 (no incluirlos):"],
    ["   - registro_sanitario_bolivia (universalidad)"],
    ["   - calibracion_proxima (gestionado vía MP/OT)"],
    ["   - responsable_username (asignación flexible en OT/MP)"],
    [""],
    ["9. Puede eliminar las filas de ejemplo y usar sus propios datos."],
    ["   Mantenga los encabezados de columna en la fila 1."],
]

# Crear Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Equipos CMMS-BioAI"

# Encabezados
ws.append(EQ_HEADERS)
for col_num, _ in enumerate(EQ_HEADERS, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = THIN_BORDER

# Datos
for fila in EQ_DATOS:
    ws.append(fila)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(EQ_HEADERS)):
    for cell in row:
        cell.border = THIN_BORDER
        cell.alignment = DATA_ALIGN

# Anchos
for i, ancho in enumerate(EQ_ANCHOS, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
ws.row_dimensions[1].height = 32

# Hoja de instrucciones
ws2 = wb.create_sheet("Instrucciones")
for row in EQ_INSTR:
    ws2.append(row)
ws2.column_dimensions['A'].width = 55
ws2.column_dimensions['B'].width = 60

# Metadata
wb.properties.creator = "CMMS-BioAI"
wb.properties.title = "Plantilla Equipos v0.9.0"

xlsx_path = OUTPUT_DIR / "plantilla_equipos.xlsx"
wb.save(xlsx_path)
wb.close()
print(f"  ✅ {xlsx_path.name} ({xlsx_path.stat().st_size} bytes)")

# Crear CSV también
csv_path = OUTPUT_DIR / "plantilla_equipos.csv"
with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(EQ_HEADERS)
    for fila in EQ_DATOS:
        writer.writerow(fila)
print(f"  ✅ {csv_path.name} ({csv_path.stat().st_size} bytes)")

print("\n✅ Plantilla de equipos v0.9.0 generada correctamente")
